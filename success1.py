import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import cv2
import pytesseract
import re

# Global variables
index = 1
folder_path = 'zource'
last_amt = []
processed_files = set()

def pdf_to_image(pdf_folder_path):    

    pdf_files = [file for file in os.listdir(pdf_folder_path) if file.endswith('.pdf')]
    
    # Convert each PDF file to images
    for pdf_file in pdf_files:
        # Construct the file paths
        pdf_file_path = os.path.join(pdf_folder_path, pdf_file)
        image_file_path = os.path.join(pdf_folder_path, os.path.splitext(pdf_file)[0] + '.png')
    
        # Convert PDF to list of PIL images
        images = convert_from_path(pdf_file_path)
    
        # Save each page of the PDF as an image file
        for i, image in enumerate(images):
            image.save(image_file_path, 'PNG')
    
    print('PDFs converted to images successfully.')

# Function to extract text from an image using Tesseract OCR
def extract_image_to_text(image_path):
    image = cv2.imread(image_path)
    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    extracted_text = pytesseract.image_to_string(gray_image)
    return extracted_text

#regEx patterns
def details_regEx_patterns():
    global bank_found, dates_found, amounts_found, payer_name_found, cuit_found, proof_number_found

    bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
    
    dates_found = re.findall(date_pattern, extracted_text)
    amounts_found = re.findall(amount_pattern, extracted_text)
    payer_name_found = re.findall(payer_name_pattern , extracted_text)
    cuit_found = re.findall(cuit_pattern, extracted_text)
    proof_number_found = re.findall(proof_number_pattern, extracted_text)

    return bank_found, dates_found, amounts_found, payer_name_found, cuit_found, proof_number_found

# Extract the first date, amount, and CUIT number if any are found
def extract_details():
    global bank, date, amount, payer, cuit, proof_number
    
    bank = bank_found[0] if bank_found else None
    date = dates_found[0] if dates_found else None
    amount = amounts_found[0] if amounts_found else None
    proof_number = proof_number_found[0] if proof_number_found else None
    
    payer = payer_name_found[0] if payer_name_found else None
    cuit = cuit_found[0] if cuit_found else None

#return the extracted details
def get_extracted_details(bank, date, amount, payer, cuit, proof_number):
    return {
        'BANCO': bank,
        'FECHA': date,
        'IMPORTE': amount,
        'TITULAR': payer,
        'CUIT': cuit,
        'NRO COMPROBANTE': proof_number
    }

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Extracted Data"

# Add headers to Excel sheet
headers = ['Serie', 'FECHA', 'IMPORTE', 'NRO COMPROBANTE', 'TITULAR', 'CUIT', 'BANCO']
ws.append(headers)

# Apply styles to header cells
bold_font = Font(bold=True)
for cell in ws[1]:
    cell.font = bold_font

# Apply padding to cells
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Function to check images in the folder and add data to Excel sheet
def check_image_and_padding(folder_path):
    global index, last_amt
     
    # total_sum_formula = Decimal(0)

     # Start from the second row (after the header row)
     # Process each image in the folder

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for filename in os.listdir(folder_path):
        if filename.endswith(".jpg") or filename.endswith(".png") or filename.endswith(".jpeg"):
            # Extract data from the image
            image_path = os.path.join(folder_path, filename)
            extracted_data = extract_data_from_image(image_path)

            # print(f"Processing image: {filename}")
            # print(f"Extracted data: {extracted_data}")

            row_data = [index] # Add the serial number as the first element
            row_data.extend([extracted_data.get(header) for header in headers[1:]])
            ws.append(row_data)
            index += 1            
            # row_data = [index] + [extracted_data.get(header) for header in headers[1:]]

            # Update total sum formula dynamically
            if extracted_data['IMPORTE']:
                amt = re.sub(r'[^\$\s0-9.]', '', extracted_data['IMPORTE'])
                num_amt = Decimal(re.sub(r'[^\d.]', '', amt))
                # total_sum_formula += num_amt

                last_amt.append(num_amt)
                print(f"Current total sum: {sum(last_amt)}")
            
            # Apply yellow fill to cells that are empty in the row
            if any(is_empty(cell.value) for cell in ws[ws.max_row]):
                for cell in ws[ws.max_row]:
                    if is_empty(cell.value):
                        cell.fill = yellow_fill

    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Adjust the multiplier as needed
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

# Function to check if a cell is empty
def is_empty(cell_value):
    return cell_value is None or cell_value == ''

# Function to extract data from an image based on its content
def extract_data_from_image(image_path):
    global index
    extracted_text = extract_image_to_text(image_path)
    lines = extracted_text.split('\n')

    bank_name = None
    for line in lines:
        if "bancopatagonia" in line.lower():
            bank_name = "Bancopatagonia"
            break
        elif "galicia" in line.lower():
            bank_name = "Galicia"
            break
        elif "mercado pago" in line.lower():
            bank_name = "Mercado pago"
            break
        # elif "santander" in line.lower():
        #     bank_name = "Santander"
        #     break

        #'fo al' because it is not dectecting the CUenta DNI
        # elif "cuentadni" in line.lower():
        #     bank_name = "Cuenta DNI"
        #     break

        # hold this becauze it has the 'Santander' word in the middle of the text
        elif "bna" in line.lower():
            # print(line.lower())
            bank_name = "BNA"
            break

        # hold this becauze it has the 'Santander' word in the middle of the text
        elif "supervielle" in line.lower():
            # print(line.lower())
            bank_name = "SUPERVIELLE"
            break

        elif "bancociudad" in line.lower():
            # print(line.lower())
            bank_name = "BancoCiudad"
            break

        elif "banco santa fe" in line.lower():
            # print(line.lower())
            bank_name = "Banco Santa Fe"
            break

        elif "bbva" in line.lower():
            # print(line.lower())
            bank_name = "BBVA"
            break

        elif "naranja x" in line.lower():
            # print(line.lower())
            bank_name = "Naranja X"
            break

        elif "banco credicoop coop. ltdo" in line.lower():
            # print(line.lower())
            bank_name = "Banco Credicoop Coop. Ltdo"
            break

        elif "personal pay" in line.lower():
            # print(line.lower())
            bank_name = "Personal Pay"
            break

        elif "bancor" in line.lower():
            bank_name = "Bancor"
            break

        elif "xp" in line.lower(): # or 'ars' coz #HSBC is not detected correctly either "xp" as bank symbol and "ARS" as currency
            print(line.lower())
            bank_name = "HSBC"
            break

        elif "uala" in line.lower(): 
            # print(line.lower())
            bank_name = "Uala" 
            break

    if bank_name == "Bancopatagonia":
        global bank, date, amount, payer, cuit, proof_number
        bank_pattern = 'Bancopatagonia'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{10}\b'
        payer_name_pattern = lines[11]
        cuit_pattern = 'None'  # MARTIN SAID Cuit IS FILLED MANUALLY

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        bank = bank_found[0] if bank_found else None
        date = dates_found[0] if dates_found else None
        amount = amounts_found[1] if amounts_found else None
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "Galicia":
        bank_pattern = 'Galicia'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{11}\b'
        payer_name_pattern = lines[8]
        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        bank = bank_found[0] if bank_found else None
        date = dates_found[0] if dates_found else None
        amount = amounts_found[0] if amounts_found else None
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None

        if any(len(p) == 9 for p in proof_number_found) and any(len(p) == 11 for p in proof_number_found):
            proof_number = next((p for p in proof_number_found if len(p) == 9), None)
        else:
            proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "Mercado pago":
        bank_pattern = 'Mercado pago'
        date_pattern = r'\b\d{1,2} de [a-z]+ \d{4}\b'
        amount_pattern = r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{11}\b'
        payer_name_pattern = r'(?:de )?([A-Z][a-z]+(?: [A-Z][a-z]+)*)'
        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        bank = bank_found[0] if bank_found else None
        date = dates_found[0] if dates_found else None
        amount = amounts_found[0] if amounts_found else None
        payer = payer_name_found[1] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

        """elif bank_name == "Santander":
        bank_pattern = 'Santander'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = 'None' #MARTIN SAID NAME IS FILLED MANUALLY 
        cuit_pattern = 'None' #MARTIN SAID Cuit IS FILLED MANUALLY 

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        bank = bank_found[0] if bank_found else None
        date = dates_found[0] if dates_found else None
        amount = amounts_found[0] if amounts_found else None
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None"""

        """elif bank_name == "Cuenta DNI":
        lines_7 = lines[7]
        bank_pattern = 'Cuenta DNI'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{6}\b'
        payer_name_pattern = lines_7  
        cuit_pattern = r'\b\d{11}\b'

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        # bank = bank_found[0] if bank_found else None
        bank = bank_pattern
        date = dates_found[0] if dates_found else None
        amount = amounts_found[0] if amounts_found else None
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None"""

   # hold this becauze it has the 'Santander' word in the middle of the text
    elif bank_name == "BNA":
        bank_pattern = 'BNA'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = 'None' #MARTIN SAID NAME IS FILLED MANUALLY 
        cuit_pattern = r'\b\d{11}\b' #MARTIN SAID Cuit IS FILLED MANUALLY 

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        bank = bank_found[0] if bank_found else None
        date = dates_found[0] if dates_found else None
        amount = amounts_found[0] if amounts_found else None
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None 

    # hold this becauze it has the 'Santander' word in the middle of the text
    elif bank_name == "SUPERVIELLE":
        bank_pattern = 'SUPERVIELLE'
        date_pattern = r'\b\d{1,2}\s+[A-Za-z]+\s+\d{4}\b'
        amount_pattern = r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{4}\b'
        payer_name_pattern = 'None' #MARTIN SAID NAME IS FILLED MANUALLY 
        cuit_pattern = 'None' #MARTIN SAID Cuit IS FILLED MANUALLY 

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        bank = bank_found[0] if bank_found else None
        date = dates_found[0] if dates_found else None
        amount = amounts_found[0] if amounts_found else None
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "BancoCiudad":

        lines_33 = lines[33] + lines[34]

        bank_pattern = 'BancoCiudad'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = lines_33
        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        bank = bank_found[0] if bank_found else None
        date = dates_found[0] if dates_found else None
        amount = amounts_found[0] if amounts_found else None
        payer = re.sub(r'[0-9,-]+', '', payer_name_pattern)
        cuit = cuit_found[1] if cuit_found else None
        proof_number = proof_number_found[1] if proof_number_found else None

    elif bank_name == "Banco Santa Fe":

        line_24 = lines[24] 

        bank_pattern = 'Banco Santa Fe'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = line_24
        cuit_pattern = r'\b\d{11}\b'

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        bank = bank_found[0] if bank_found else None
        date = dates_found[0] if dates_found else None
        amount = amounts_found[0] if amounts_found else None
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "BBVA":

        #for payer name
        line_4 = lines[4]
        line_4 = re.sub(r'^\w+:+\s+\b', '', line_4)
        line_4 = re.sub(r'\b\s+\w+$', '', line_4)

        #for proof number
        line_3 = lines[3]
        Cuenta_Origen = r'^Cuenta Origen:\s+CC\s+\$\s+\d{4}-\d{6}\/\d{1}\s+'
        line_3 = re.sub(Cuenta_Origen, '', line_3)
        line_3 = re.sub(r' ' , '', line_3)

        #for amount
        line_12 = lines[12]
        #removes 1st word
        line_12 = re.sub(r'^\w+:+\s+\b', '', line_12)
        #adds $ 
        line_12 = '$ ' + line_12
        line_12 = re.sub(r'\b,00+$', '', line_12)


        bank_pattern = 'Banco Santa Fe'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        #something is wrong here
        amount_pattern = line_12
        payer_name_pattern = line_4
        proof_number_pattern = line_3

        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        bank = bank_found[0] if bank_found else None
        date = dates_found[0] if dates_found else None
        amount = line_12
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = line_3

    elif bank_name == "Naranja X":

        #for date
        line_8 = lines[8]
        line_8 = re.sub(r'^\w+\s+\w+\s+\b|\b\s-\s\d{2}:\d{2}\s+h$', '', line_8)
        #for extracting payer name
        line_14 = lines[14] 


        bank_pattern = 'Naranja X'
        date_pattern = line_8
        amount_pattern = r'\$\s*\d+\.?\d*'
        payer_name_pattern = line_14
        proof_number_pattern = r'\b\d{10}\b'

        cuit_pattern = r'\b\d{11}\b'

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        bank = bank_found[0] if bank_found else None
        date = dates_found[0] if dates_found else None
        amount = amounts_found[0] if amounts_found else None
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "Banco Credicoop Coop. Ltdo":

        # #for bank name if in case it is not detected uncomment below
        # line_2 = lines[2] #CREDICOOP Banco Credicoop Coop. Ltdo
        # line_2 = re.sub(r'^\w+\s+\b', '', line_2)


        bank_pattern = 'Banco Credicoop Coop. Ltdo' #line_2
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        payer_name_pattern = 'None'
        proof_number_pattern = r'\b\d{9}\b'

        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        bank = bank_found[0] if bank_found else None
        date = dates_found[0] if dates_found else None
        amount = amounts_found[1] if amounts_found else None
        amount = amount.replace('$', '$ ')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "Personal Pay":

        #if the amount has 1 in it, the OCR is dectecting it as ] that's why we have to check the first line
        if lines[0] == 'personal pay':

            #for amount
            line_4 = lines[4]
            line_4 = re.sub(r'"', '', line_4) # remove '' in the end of the amount
            line_4 = re.sub(r'(?<=\$)(?=\d)', r' ', line_4) # adds '$ ' in place of '$'

            line_39 = lines[39] #43 for first proof number
            line_40 = line_39 + lines[40] #44 for second proof number

            #for payer name
            line_28 = lines[28]

        bank_pattern = 'Personal Pay'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = line_4
        payer_name_pattern = line_28
        proof_number_pattern = line_40

        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        bank = bank_found[0] if bank_found else None
        date = dates_found[0] if dates_found else None
        amount = line_4
        payer = line_28
        cuit = cuit_found[0] if cuit_found else None
        proof_number = line_40

    elif bank_name == "Bancor":

        #for proof number
        line_3 = lines[3] 
        #for payer name
        line_11 = lines[11]
        line_11 = re.sub(r'^\w+:+\s+\b', '', line_11)

        bank_pattern = 'Bancor'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = line_3
        payer_name_pattern = line_11
        cuit_pattern = r'\b\d{11}\b'  # MARTIN SAID Cuit IS FILLED MANUALLY

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        bank = bank_found[0] if bank_found else None
        date = dates_found[0] if dates_found else None
        amount = amounts_found[1] if amounts_found else None
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "HSBC":

        #for payer name
        line_11 = lines[11]
        line_11 = re.sub(r'^\w+:+\s+\b', '', line_11)

        bank_pattern = 'Uala' #'usec' #HSBC is detected as that
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern =  r'\bARS\s*\d+\.?\d*'  # have to Replace ARS with $
        proof_number_pattern = r'\b\d{4}\b'
        payer_name_pattern = "None"
        cuit_pattern = "None"  # MARTIN SAID Cuit IS FILLED MANUALLY

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        bank = "Uala"
        date = dates_found[0] if dates_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('ARS', '$ ') #replaces '$' with '$ '
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[1] if proof_number_found else None

    elif bank_name == "Uala":

        # for  uala.png and uala1.png formats
        if lines[0] == "afr":
            # global line_8, line_10
            #payer name
            line_8 = lines[9]
            line_8 = line_8.replace("Nombre remitente ", "")
            # print(line_8)
            #for proof number
            line_10 = lines[11]
            line_10 = line_10.replace("Id Op. ", "")
            # print(line_10)
        if lines[0] == "VAD Comprobante de transferencia":
            #payer name
            line_8 = lines[8]
            line_8 = line_8.replace("Nombre remitente ", "")
            # print(line_8)
            #for proof number
            line_10 = lines[10]
            line_10 = line_10.replace("Id Op. ", "")

        bank_pattern = 'Uala'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = line_10
        payer_name_pattern = line_8
        cuit_pattern = "None"  # MARTIN SAID Cuit IS FILLED MANUALLY

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        bank = bank_found[0] if bank_found else None
        date = dates_found[0] if dates_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('$', '$ ') #replaces '$' with '$ '
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    # Return the extracted data
    return {
        'BANCO': bank_name,
        'FECHA': date,
        'IMPORTE': amount,
        'NRO COMPROBANTE': proof_number,
        'TITULAR': payer,
        'CUIT': cuit
    }


### heres a bit catchy 

# sum and adding and saving and opening the file
print(f"total sum: {last_amt}")
print(f"Current total sum: {sum(last_amt)}")

total_sum = sum(last_amt)
# Append the total sum row to the Excel sheet
total_sum_row = ['TOTAL', None, f'${total_sum}', None, None, None, None]
ws.append(total_sum_row)

from datetime import date

today = date.today()
formatted_date = today.strftime("%d_%m_%Y")

# Specify the path to the extracted file
extracted_file_path = f'{formatted_date}_extracted_info.xlsx'


# Save the Excel file
wb.save(extracted_file_path)
import subprocess
import os

# Open the file with LibreOffice in linux
subprocess.run(['libreoffice', extracted_file_path])

# Open the file with Microsoft Excel in windows
#os.system("start EXCEL.EXE extracted_info.xlsx")