#libraries needed
import cv2
import pytesseract
from pdf2image import convert_from_path
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from decimal import Decimal
from openpyxl.styles import PatternFill
from datetime import date


#global variable for serial number
index = 1
last_amt = [] 

def pdf_to_image(pdf_folder_path):    

    pdf_files = [file for file in os.listdir(pdf_folder_path) if file.endswith('.pdf')]
    
    # Convert each PDF file to images
    for pdf_file in pdf_files:
        # Construct the file paths
        pdf_file_path = os.path.join(pdf_folder_path, pdf_file)
        image_file_path = os.path.join(pdf_folder_path, os.path.splitext(pdf_file)[0] + '.png')
    
        # Convert PDF to list of PIL images
        images = convert_from_path(pdf_file_path)
    
        # Save each page of the PDF as an image file
        for i, image in enumerate(images):
            image.save(image_file_path, 'PNG')
    
    print('PDFs converted to images successfully.')

#extract the image to text
def extract_image_to_text(image_path):    
    global extracted_text

    image = cv2.imread(image_path)
    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    extracted_text = pytesseract.image_to_string(gray_image)

#regEx patterns
def details_regEx_patterns():
    global bank_found, dates_found, amounts_found, payer_name_found, cuit_found, proof_number_found

    bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
    dates_found = re.findall(date_pattern, extracted_text)
    amounts_found = re.findall(amount_pattern, extracted_text)
    payer_name_found = re.findall(payer_name_pattern , extracted_text)
    cuit_found = re.findall(cuit_pattern, extracted_text)
    proof_number_found = re.findall(proof_number_pattern, extracted_text)

    return bank_found, dates_found, amounts_found, payer_name_found, cuit_found, proof_number_found

# Extract the first date, amount, and CUIT number if any are found
def extract_details():
    global bank, date, amount, payer, cuit, proof_number
    
    bank = bank_found[0] if bank_found else None
    date = dates_found[0] if dates_found else None
    amount = amounts_found[0] if amounts_found else None
    proof_number = proof_number_found[0] if proof_number_found else None
    payer = payer_name_found[0] if payer_name_found else None
    cuit = cuit_found[0] if cuit_found else None

#return the extracted details
def get_extracted_details(bank, date, amount, payer, cuit, proof_number):
    return {
        'BANCO': bank,
        'FECHA': date,
        'IMPORTE': amount,
        'TITULAR': payer,
        'CUIT': cuit,
        'NRO COMPROBANTE': proof_number
    }

wb = Workbook()
ws = wb.active
ws.title = "Extracted Data"


headers = [ 'Serie', 'FECHA', 'IMPORTE', 'NRO COMPROBANTE', 'TITULAR', 'CUIT', 'BANCO']
ws.append(headers)

bold_font = Font(bold=True)
for cell in ws[1]:
    cell.font = bold_font

for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def check_image_and_padding(folder_path):
    global index, last_amt
     
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for filename in os.listdir(folder_path):
        if filename.endswith(".jpg") or filename.endswith(".png") or filename.endswith(".jpeg"):

            image_path = os.path.join(folder_path, filename)
            extracted_data = extract_data_from_image(image_path)

            row_data = [index] 
            row_data.extend([extracted_data.get(header) for header in headers[1:]])
            ws.append(row_data)
            index += 1            

            if extracted_data['IMPORTE']:
                amt = re.sub(r'[^\$\s0-9.]', '', extracted_data['IMPORTE'])
                num_amt = Decimal(re.sub(r'[^\d.]', '', amt))
                # total_sum_formula += num_amt

                last_amt.append(num_amt)
            
            if any(is_empty(cell.value) for cell in ws[ws.max_row]):
                for cell in ws[ws.max_row]:
                    if is_empty(cell.value):
                        cell.fill = yellow_fill
    
    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

def is_empty(cell_value):
    return cell_value is None or cell_value == ''

#bancopatagonia
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    pdf_folder_path = 'bancopatagonia'
    pdf_to_image(pdf_folder_path)
    extract_image_to_text(image_path)
    lines = extracted_text.split('\n')
    lines_12 = lines[11]

    bank_pattern = 'bancopatagonia'
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{10}\b'
    payer_name_pattern = lines_12
    cuit_pattern =  'None' #MARTIN said Cuit IS FILLED MANUALLY 

    details_regEx_patterns()

    extract_details()

    amount = amounts_found[1] if amounts_found else None

    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)


folder_path = 'bancopatagonia'
check_image_and_padding(folder_path)

#galicia
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    pdf_folder_path = 'galicia'
    pdf_to_image(pdf_folder_path)
    extract_image_to_text(image_path)
    lines = extracted_text.strip().split('\n')
    line_9 = lines[8]

    bank_pattern = 'galicia'
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{9,11}\b'
    payer_name_pattern = line_9 
    cuit_pattern =  r'\b\d{2}-\d{8}-\d{1}\b'

    details_regEx_patterns()

    extract_details()

    if any(len(p) == 9 for p in proof_number_found) and any(len(p) == 11 for p in proof_number_found):
        proof_number = next((p for p in proof_number_found if len(p) == 9), None)
    else:
        proof_number = proof_number_found[0] if proof_number_found else None
        
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

folder_path = 'galicia'
check_image_and_padding(folder_path)


#mercado pago
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    pdf_folder_path = 'mercado_pago'
    pdf_to_image(pdf_folder_path)
    extract_image_to_text(image_path)

    bank_pattern = 'mercado pago'
    date_pattern = r'\b\d{1,2} de [a-z]+ \d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*'
    payer_name_pattern = r'(?:de )?([A-Z][a-z]+(?: [A-Z][a-z]+)*)'
    cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'
    proof_number_pattern = r'\b\d{11}\b'

    details_regEx_patterns()
    extract_details()
    payer = payer_name_found[1] if payer_name_found else None

    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

folder_path = 'mercado_pago'
check_image_and_padding(folder_path)

#santander
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern
    
    pdf_folder_path = 'santander'
    pdf_to_image(pdf_folder_path)
    extract_image_to_text(image_path)

    bank_pattern = 'santander'
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{8}\b'
    payer_name_pattern = 'None' #MARTIN SAID NAME IS FILLED MANUALLY 
    cuit_pattern =  'None' #MARTIN SAID Cuit IS FILLED MANUALLY 

    details_regEx_patterns()
    extract_details()
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)
folder_path = 'santander'
check_image_and_padding(folder_path)


#supervielle
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern
    
    pdf_folder_path = 'supervielle'
    pdf_to_image(pdf_folder_path)
    extract_image_to_text(image_path)

    bank_pattern = 'supervielle'
    date_pattern = r'\b\d{1,2}\s+[A-Za-z]+\s+\d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{4}\b' 
    payer_name_pattern = 'None' #MARTIN said NAME IS FILLED MANUALLY 
    cuit_pattern =  'None' #MARTIN said Cuit IS FILLED MANUALLY 

    details_regEx_patterns()
    extract_details()
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

folder_path = 'supervielle'
check_image_and_padding(folder_path)

#bna
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern
    extract_image_to_text(image_path)

    bank_pattern = 'bna'
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{8}\b' 
    payer_name_pattern = 'None' #MARTIN said NAME IS FILLED MANUALLY 
    cuit_pattern =  r'\b\d{11}\b'  #MARTIN said Cuit IS FILLED MANUALLY 

    details_regEx_patterns()
    extract_details()
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

folder_path = 'bna'
check_image_and_padding(folder_path)

#cuenta_dni  # bank image is not retriving but as of now okay coz folders are used for bank name
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern
    extract_image_to_text(image_path)

    lines = extracted_text.split('\n')
    lines_8 = lines[7]

    bank_pattern = 'Cuenta DNI'
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{6}\b' 
    payer_name_pattern = lines_8
    cuit_pattern =  r'\b\d{11}\b' 

    details_regEx_patterns()
    extract_details()
    bank = bank_pattern
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)
folder_path = 'cuenta_dni'
check_image_and_padding(folder_path)


#Banco Ciudad
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    pdf_folder_path = 'banco_ciudad'
    pdf_to_image(pdf_folder_path)

    extract_image_to_text(image_path)

    lines = extracted_text.strip().split('\n')
    line_33 = lines[33] 
    line_34 = lines[34]
    line = line_33 + line_34

    bank_pattern = 'BancoCiudad'
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{8}\b'
    payer_name_pattern = line 
    cuit_pattern =  r'\b\d{2}-\d{8}-\d{1}\b'

    details_regEx_patterns()
    extract_details()
    proof_number = proof_number_found[1] if proof_number_found else None
    payer = re.sub(r'[0-9,-]+', '', payer_name_pattern)
    cuit = cuit_found[1] if cuit_found else None
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

folder_path = 'banco_ciudad'
check_image_and_padding(folder_path)

#Banco Santa Fe
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    pdf_folder_path = 'banco_santa_fe'
    pdf_to_image(pdf_folder_path)
    extract_image_to_text(image_path)
    lines = extracted_text.strip().split('\n')
    line_24 = lines[24] 
    line = line_24

    bank_pattern = 'Banco Santa Fe'
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{8}\b'
    payer_name_pattern = line 
    cuit_pattern =  r'\b\d{11}\b'

    details_regEx_patterns()
    extract_details()
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

folder_path = 'banco_santa_fe'
check_image_and_padding(folder_path)

#BBVA 
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    pdf_folder_path = 'bbva'
    pdf_to_image(pdf_folder_path)
    extract_image_to_text(image_path)
    lines = extracted_text.strip().split('\n')

    line_4 = lines[4]
    line_4 = re.sub(r'^\w+:+\s+\b', '', line_4)
    line_4 = re.sub(r'\b\s+\w+$', '', line_4)
    line_3 = lines[3]
    Cuenta_Origen = r'^Cuenta Origen:\s+CC\s+\$\s+\d{4}-\d{6}\/\d{1}\s+'
    line_3 = re.sub(Cuenta_Origen, '', line_3)
    line_3 = re.sub(r' ' , '', line_3)
    line_12 = lines[12]
    line_12 = re.sub(r'^\w+:+\s+\b', '', line_12)
    line_12 = '$ ' + line_12
    line_12 = re.sub(r'\b,00+$', '', line_12)

    bank_pattern = 'BBVA'
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    amount_pattern = line_3
    payer_name_pattern = line_4
    proof_number_pattern = line_12
    cuit_pattern =  r'\b\d{2}-\d{8}-\d{1}\b'


    details_regEx_patterns()
    extract_details()
    amount = line_12
    proof_number = line_3
    payer = payer_name_found[0] if payer_name_found else None 
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)
folder_path = 'bbva'
check_image_and_padding(folder_path)

#Naranja X
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    pdf_folder_path = 'naranja_x'
    pdf_to_image(pdf_folder_path)
    extract_image_to_text(image_path)
    lines = extracted_text.strip().split('\n')

    line_8 = lines[8]
    line_8 = re.sub(r'^\w+\s+\w+\s+\b|\b\s-\s\d{2}:\d{2}\s+h$', '', line_8)
    line_14 = lines[14] 

    bank_pattern = 'naranja x'
    date_pattern = line_8 
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{10}\b'
    payer_name_pattern = line_14
    cuit_pattern =  r'\b\d{11}\b'

    details_regEx_patterns()
    extract_details()
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)
folder_path = 'naranja_x'
check_image_and_padding(folder_path)


total_sum = sum(last_amt)
total_sum_row = ['TOTAL', None, f'${total_sum}', None, None, None, None]
ws.append(total_sum_row)

from datetime import date

today = date.today()
formatted_date = today.strftime("%d_%m_%Y")

extracted_file_path = f'{formatted_date}_extracted_info.xlsx'
wb.save(extracted_file_path)

import subprocess
import os

# # Open the file automatically with LibreOffice in linux
# subprocess.run(['libreoffice', extracted_file_path])

# sorry i don't know about MAC os ;)

# Opens the file automatically with Microsoft Excel in windows
os.system("start EXCEL.EXE extracted_info.xlsx")
