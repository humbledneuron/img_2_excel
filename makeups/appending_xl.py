import os
from openpyxl import load_workbook

# Create a new Excel workbook for consolidation
consolidated_wb = Workbook()
consolidated_ws = consolidated_wb.active
consolidated_ws.title = "Consolidated Data"

# Define the folder path containing the Excel files to be consolidated
folder_path = 'folder_with_xlsx_files'

# Iterate through each file in the folder
for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx"):
        # Load the Excel file
        file_path = os.path.join(folder_path, filename)
        source_wb = load_workbook(file_path)
        source_ws = source_wb.active

        # Iterate through rows in the source worksheet and append them to the consolidated worksheet
        for row in source_ws.iter_rows(min_row=2):  # Start from the second row to skip headers
            values = [cell.value for cell in row]
            consolidated_ws.append(values)

# Save the consolidated workbook
consolidated_wb.save('consolidated_data.xlsx')
