#libraries needed
import cv2
import pytesseract
from pdf2image import convert_from_path
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from decimal import Decimal
from openpyxl.styles import PatternFill
from datetime import date


#global variable for serial number
index = 1
last_amt = [] #list


# Get the list of PDF files in the folder
# def pdf_to_image(pdf_folder_path):    
def pdf_to_image(pdf_folder_path):    
    # pdf_folder_path = 'pdf_source'
    # source_folder_path = 'source'

    pdf_files = [file for file in os.listdir(pdf_folder_path) if file.endswith('.pdf')]
    
    # Convert each PDF file to images
    for pdf_file in pdf_files:
        # Construct the file paths
        pdf_file_path = os.path.join(pdf_folder_path, pdf_file)
        image_file_path = os.path.join(pdf_folder_path, os.path.splitext(pdf_file)[0] + '.png')
    
        # Convert PDF to list of PIL images
        images = convert_from_path(pdf_file_path)
    
        # Save each page of the PDF as an image file
        for i, image in enumerate(images):
            image.save(image_file_path, 'PNG')
    
    print('PDFs converted to images successfully.')

#extract the image to text
def extract_image_to_text(image_path):    
    global extracted_text

    image = cv2.imread(image_path)
    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    extracted_text = pytesseract.image_to_string(gray_image)

#regEx patterns
def details_regEx_patterns():
    global bank_found, dates_found, amounts_found, payer_name_found, cuit_found, proof_number_found

    bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
    
    dates_found = re.findall(date_pattern, extracted_text)
    amounts_found = re.findall(amount_pattern, extracted_text)
    payer_name_found = re.findall(payer_name_pattern , extracted_text)
    cuit_found = re.findall(cuit_pattern, extracted_text)
    proof_number_found = re.findall(proof_number_pattern, extracted_text)

    return bank_found, dates_found, amounts_found, payer_name_found, cuit_found, proof_number_found

# Extract the first date, amount, and CUIT number if any are found
def extract_details():
    global bank, date, amount, payer, cuit, proof_number
    
    bank = bank_found[0] if bank_found else None
    date = dates_found[0] if dates_found else None
    amount = amounts_found[0] if amounts_found else None
    proof_number = proof_number_found[0] if proof_number_found else None
    
    payer = payer_name_found[0] if payer_name_found else None
    cuit = cuit_found[0] if cuit_found else None

#return the extracted details
def get_extracted_details(bank, date, amount, payer, cuit, proof_number):
    return {
        'BANCO': bank,
        'FECHA': date,
        'IMPORTE': amount,
        'TITULAR': payer,
        'CUIT': cuit,
        'NRO COMPROBANTE': proof_number
    }

#new Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Extracted Data"

# Add headers to the Excel sheet
headers = [ 'Serie', 'FECHA', 'IMPORTE', 'NRO COMPROBANTE', 'TITULAR', 'CUIT', 'BANCO']
ws.append(headers)

# Apply bold style to the header cells
bold_font = Font(bold=True)
for cell in ws[1]:
    cell.font = bold_font

# Apply padding to all cells
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

#check if image exists and adjust the padding

def check_image_and_padding(folder_path):
    global index, last_amt
     
    # total_sum_formula = Decimal(0)

     # Start from the second row (after the header row)
     # Process each image in the folder

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for filename in os.listdir(folder_path):
        if filename.endswith(".jpg") or filename.endswith(".png") or filename.endswith(".jpeg"):
            # Extract data from the image
            image_path = os.path.join(folder_path, filename)
            extracted_data = extract_data_from_image(image_path)

            # print(f"Processing image: {filename}")
            # print(f"Extracted data: {extracted_data}")

            row_data = [index] # Add the serial number as the first element
            row_data.extend([extracted_data.get(header) for header in headers[1:]])
            ws.append(row_data)
            index += 1            
            # row_data = [index] + [extracted_data.get(header) for header in headers[1:]]

            # Update total sum formula dynamically
            if extracted_data['IMPORTE']:
                amt = re.sub(r'[^\$\s0-9.]', '', extracted_data['IMPORTE'])
                num_amt = Decimal(re.sub(r'[^\d.]', '', amt))
                # total_sum_formula += num_amt

                last_amt.append(num_amt)
                print(f"Current total sum: {sum(last_amt)}")
            
            # Apply yellow fill to cells that are empty in the row
            if any(is_empty(cell.value) for cell in ws[ws.max_row]):
                for cell in ws[ws.max_row]:
                    if is_empty(cell.value):
                        cell.fill = yellow_fill
    
    #         print(f"Current total sum: {total_sum_formula}")

    # # Append the total sum row to the Excel sheet
    # total_sum_row = ['TOTAL', f'${total_sum_formula}', None, None, None, None]
    # ws.append(total_sum_row)
    
    # Automatically adjust column widths based on content
    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Adjust the multiplier as needed
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

# Function to check if a cell is empty
def is_empty(cell_value):
    return cell_value is None or cell_value == ''

#bancopatagonia
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    pdf_folder_path = 'bancopatagonia/bancopatagonia_assets'
    pdf_to_image(pdf_folder_path)
    extract_image_to_text(image_path)
    lines = extracted_text.split('\n')
    lines_11 = lines[11]
    # # regEx patterns for different types of data to extract
    bank_pattern = 'bancopatagonia'
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{10}\b'
    payer_name_pattern = lines_11
    cuit_pattern =  'None' #MARTIN said Cuit IS FILLED MANUALLY 
    # Extract information using the regEx patterns
    details_regEx_patterns()
    # Extract the date, amount, and CUIT number if any are found
    extract_details()
    #this is important to extract the [1] amount coz the fun has [0] default
    amount = amounts_found[1] if amounts_found else None
    # Return the extracted details
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

# Define the folder containing the images
folder_path = 'bancopatagonia/bancopatagonia_assets'
#function 
check_image_and_padding(folder_path)

#galicia
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    pdf_folder_path = 'galicia/galicia_assets'
    pdf_to_image(pdf_folder_path)

    extract_image_to_text(image_path)

    lines = extracted_text.strip().split('\n')
    line_8 = lines[8]

    # Define regular expression patterns for different types of data to extract

    bank_pattern = 'galicia'
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{9,11}\b'
    
    payer_name_pattern = line_8
    
    cuit_pattern =  r'\b\d{2}-\d{8}-\d{1}\b'


    # Extract information using the regEx patterns
    details_regEx_patterns()

    # Extract the date, amount, and CUIT number if any are found
    extract_details()

    # Check if both 9-digit and 11-digit proof numbers are present
    if any(len(p) == 9 for p in proof_number_found) and any(len(p) == 11 for p in proof_number_found):
        # Extract the 9-digit proof number if both are present
        proof_number = next((p for p in proof_number_found if len(p) == 9), None)
    else:
        # Extract the first found proof number (either 9-digit or 11-digit)
        proof_number = proof_number_found[0] if proof_number_found else None
        
    # return extracted_details
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

# Define the folder containing the images
folder_path = 'galicia/galicia_assets'

#function 
check_image_and_padding(folder_path)

#mercado pago
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    pdf_folder_path = 'mercado_pago/mercado_pago_assets'
    pdf_to_image(pdf_folder_path)

    extract_image_to_text(image_path)

    # Define regular expression patterns for different types of data to extract
    bank_pattern = 'mercado pago'
    
    date_pattern = r'\b\d{1,2} de [a-z]+ \d{4}\b'
    # amount_pattern = r'\$\s*\d+(?:[.,]\d+)?'
    amount_pattern = r'\$\s*\d+\.?\d*'
    payer_name_pattern = r'(?:de )?([A-Z][a-z]+(?: [A-Z][a-z]+)*)'
    cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'
    proof_number_pattern = r'\b\d{11}\b'

    # Find all occurrences of dates, amounts, and CUIT numbers in the extracted text
    details_regEx_patterns()

    # Extract the date, amount, and CUIT number if any are found
    extract_details()

    #this is important to extract the [1] amount coz the fun has [0] default
    payer = payer_name_found[1] if payer_name_found else None

    # Return the extracted information as a dictionary
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

# Define the folder containing the images
folder_path = 'mercado_pago/mercado_pago_assets'

#function 
check_image_and_padding(folder_path)

#santander
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern
    
    pdf_folder_path = 'santander/santander_assets'
    pdf_to_image(pdf_folder_path)

    extract_image_to_text(image_path)

    # Define regular expression patterns for different types of data to extract
    bank_pattern = 'santander'
    
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{8}\b'
    
    payer_name_pattern = 'None' #MARTIN SAID NAME IS FILLED MANUALLY 
    cuit_pattern =  'None' #MARTIN SAID Cuit IS FILLED MANUALLY 

        # Extract information using the regEx patterns
    details_regEx_patterns()

    # Extract the date, amount, and CUIT number if any are found
    extract_details()

    # Return the extracted details
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

# Define the folder containing the images
folder_path = 'santander/santander_assets'

#function 
check_image_and_padding(folder_path)

#supervielle
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern
    
    pdf_folder_path = 'supervielle/supervielle_assets'
    pdf_to_image(pdf_folder_path)

    extract_image_to_text(image_path)

    # Define regular expression patterns for different types of data to extract
    bank_pattern = 'supervielle'
    
    date_pattern = r'\b\d{1,2}\s+[A-Za-z]+\s+\d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{4}\b' #have to ask martin about this coz there's none
    
    payer_name_pattern = 'None' #MARTIN said NAME IS FILLED MANUALLY 
    cuit_pattern =  'None' #MARTIN said Cuit IS FILLED MANUALLY 

    # Extract information using the regEx patterns
    details_regEx_patterns()

    # Extract the date, amount, and CUIT number if any are found
    extract_details()

    # Return the extracted details
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

# Define the folder containing the images
folder_path = 'supervielle/supervielle_assets'

#function 
check_image_and_padding(folder_path)

#bna
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern
    extract_image_to_text(image_path)

    # Define regular expression patterns for different types of data to extract
    bank_pattern = 'bna'
    
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{8}\b'
    
    payer_name_pattern = 'None' #MARTIN said NAME IS FILLED MANUALLY 
    cuit_pattern =  r'\b\d{11}\b'  #MARTIN said Cuit IS FILLED MANUALLY 

    # Extract information using the regEx patterns
    details_regEx_patterns()

    # Extract the date, amount, and CUIT number if any are found
    extract_details()

    # Return the extracted details
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

# Define the folder containing the images
folder_path = 'bna/bna_assets'

#function 
check_image_and_padding(folder_path)

#cuenta_dni  # bank image is not retriving
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern
    extract_image_to_text(image_path)

    lines = extracted_text.split('\n')
    lines_7 = lines[7]

    # Define regular expression patterns for different types of data to extract
    bank_pattern = 'Cuenta DNI'
    
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{6}\b' #have to ask martin about this coz there's none
    
    payer_name_pattern = lines_7
    
    cuit_pattern =  r'\b\d{11}\b' 

    # Extract information using the regEx patterns
    details_regEx_patterns()

    # Extract the date, amount, and CUIT number if any are found
    extract_details()
    bank = bank_pattern

    # Return the extracted details
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

# Define the folder containing the images
folder_path = 'cuenta_dni/cuenta_dni_assets'

#function 
check_image_and_padding(folder_path)

#Banco Ciudad
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    pdf_folder_path = 'banco_ciudad/banco_ciudad_assets'
    pdf_to_image(pdf_folder_path)

    extract_image_to_text(image_path)

    lines = extracted_text.strip().split('\n')
    line_33 = lines[33] 
    line_34 = line_33 + lines[34]
    # line = line_33 + line_34


    # Define regular expression patterns for different types of data to extract

    bank_pattern = 'BancoCiudad'
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{8}\b'
    
    payer_name_pattern = line_34 
    
    cuit_pattern =  r'\b\d{2}-\d{8}-\d{1}\b'


    # Extract information using the regEx patterns
    details_regEx_patterns()

    # Extract the date, amount, and CUIT number if any are found
    extract_details()

    #this line is important to extract the proof number
    proof_number = proof_number_found[1] if proof_number_found else None
    
    #this line is important to extract the payer name
    payer = re.sub(r'[0-9,-]+', '', payer_name_pattern)
    
    #this line is important to extract the cuit number
    cuit = cuit_found[1] if cuit_found else None
    
    # return extracted_details
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

# Define the folder containing the images
folder_path = 'banco_ciudad/banco_ciudad_assets'

#function 
check_image_and_padding(folder_path)

#Banco Santa Fe
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    pdf_folder_path = 'banco_santa_fe/banco_santa_fe_assets'
    pdf_to_image(pdf_folder_path)

    extract_image_to_text(image_path)

    lines = extracted_text.strip().split('\n')
    line_24 = lines[24] 
    line = line_24


    # Define regular expression patterns for different types of data to extract

    bank_pattern = 'Banco Santa Fe'

    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{8}\b'
    
    payer_name_pattern = line 
    
    cuit_pattern =  r'\b\d{11}\b'


    # Extract information using the regEx patterns
    details_regEx_patterns()

    # Extract the date, amount, and CUIT number if any are found
    extract_details()
    
    # return extracted_details
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

# Define the folder containing the images
folder_path = 'banco_santa_fe/banco_santa_fe_assets'
#function 
check_image_and_padding(folder_path)

#BBVA  #amount comma problem
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    pdf_folder_path = 'bbva/bbva_assets'
    pdf_to_image(pdf_folder_path)

    extract_image_to_text(image_path)

    lines = extracted_text.strip().split('\n')

    #for payer name
    line_4 = lines[4]
    #removes 1st word
    line_4 = re.sub(r'^\w+:+\s+\b', '', line_4)
    # print(line_4s)
    #removes last word
    line_4 = re.sub(r'\b\s+\w+$', '', line_4)
    # print(line_4)


    #for proof number
    line_3 = lines[3]
    #removes 1st word
    # line_3 = re.sub(r'^\w+:+\s+\b', '', line_3)
    # Cuenta_Origen = r'^Cuenta Origen:\s+CC\s+\$\s'
    Cuenta_Origen = r'^Cuenta Origen:\s+CC\s+\$\s+\d{4}-\d{6}\/\d{1}\s+'
    line_3 = re.sub(Cuenta_Origen, '', line_3)
    line_3 = re.sub(r' ' , '', line_3)
    # print(line_3)

    #for amount
    line_12 = lines[12]
    # print(line_12)
    #removes 1st word
    line_12 = re.sub(r'^\w+:+\s+\b', '', line_12)
    #adds $ 
    line_12 = '$ ' + line_12
    line_12 = re.sub(r'\b,00+$', '', line_12)

    # print(line_12)

    # Define regular expression patterns for different types of data to extract
    
    bank_pattern = 'BBVA'
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'

    amount_pattern = line_3
    payer_name_pattern = line_4
    proof_number_pattern = line_12
    
    cuit_pattern =  r'\b\d{2}-\d{8}-\d{1}\b'


    # Extract information using the regEx patterns
    details_regEx_patterns()

    # Extract the date, amount, and CUIT number if any are found
    extract_details()

    amount = line_12
    proof_number = line_3
    payer = payer_name_found[0] if payer_name_found else None #this works

    
    # return extracted_details
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

# Define the folder containing the images
folder_path = 'bbva/bbva_assets'

#function 
check_image_and_padding(folder_path)

#Naranja X
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    pdf_folder_path = 'naranja_x/naranja_x_assets'
    pdf_to_image(pdf_folder_path)

    extract_image_to_text(image_path)

    lines = extracted_text.strip().split('\n')

    #for date
    line_8 = lines[8]
    line_8 = re.sub(r'^\w+\s+\w+\s+\b|\b\s-\s\d{2}:\d{2}\s+h$', '', line_8)
    #for extracting payer name
    line_14 = lines[14] 

    # Define regular expression patterns for different types of data to extract

    bank_pattern = 'naranja x'
    date_pattern = line_8 #this is regular r'\s+\d{1,2}/[A-Za-z]/\d{4}\b'
    
    amount_pattern = r'\$\s*\d+\.?\d*'
    proof_number_pattern = r'\b\d{10}\b'
    
    payer_name_pattern = line_14
    
    cuit_pattern =  r'\b\d{11}\b'


    # Extract information using the regEx patterns
    details_regEx_patterns()

    # Extract the date, amount, and CUIT number if any are found
    extract_details()
    
    # return extracted_details
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

# Define the folder containing the images
folder_path = 'naranja_x/naranja_x_assets'

#function 
check_image_and_padding(folder_path)

#banco_credicoop_coop_ltdo
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    pdf_folder_path = 'banco_credicoop_coop_ltdo/banco_credicoop_coop_ltdo_assets'
    pdf_to_image(pdf_folder_path)

    extract_image_to_text(image_path)

    lines = extracted_text.strip().split('\n')

    #for bank name
    line_2 = lines[2] #CREDICOOP Banco Credicoop Coop. Ltdo
    line_2 = re.sub(r'^\w+\s+\b', '', line_2)

    # Define regular expression patterns for different types of data to extract

    # bank_pattern = 'Banco Credicoop Coop. Ltdo.'

    bank_pattern = line_2
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'

    proof_number_pattern = r'\b\d{9}\b'

    payer_name_pattern = 'None'  

    amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
    
    cuit_pattern =  r'\b\d{2}-\d{8}-\d{1}\b'


    # Extract information using the regEx patterns
    details_regEx_patterns()

    # Extract the date, amount, and CUIT number if any are found
    extract_details()
    amount = amounts_found[1] if amounts_found else None
    amount = amount.replace('$', '$ ')

    # return extracted_details
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

# Define the folder containing the images
folder_path = 'banco_credicoop_coop_ltdo/banco_credicoop_coop_ltdo_assets'

#function 
check_image_and_padding(folder_path)

#Personal pay
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    pdf_folder_path = 'personal_pay/personal_pay_assets'
    pdf_to_image(pdf_folder_path)

    extract_image_to_text(image_path)

    lines = extracted_text.strip().split('\n')

    if lines[0] == 'personal pay':
        global line_4, line_40, line_28
        line_4 = lines[4] #24 for amount
        line_4 = re.sub(r'"', '', line_4) # remove '' in the end of the amount
        line_4 = re.sub(r'(?<=\$)(?=\d)', r' ', line_4) # adds '$ ' in place of '$'

        line_39 = lines[39] #43 for first proof number
        line_40 = line_39 + lines[40] #44 for second proof number

        #for payer name
        line_28 = lines[28]
        # print(line_30)
        
    # Define regular expression patterns for different types of data to extract

    bank_pattern = 'personal pay'
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'

    amount_pattern = line_4

    proof_number_pattern = line_40
    
    payer_name_pattern = line_28
    
    cuit_pattern =  r'\b\d{2}-\d{8}-\d{1}\b'


    # Extract information using the regEx patterns
    details_regEx_patterns()

    # Extract the date, amount, and CUIT number if any are found
    extract_details()
    
    amount = line_4
    proof_number = line_40
    payer = line_28

    # return extracted_details
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

# Define the folder containing the images
folder_path = 'personal_pay/personal_pay_assets'

#function 
check_image_and_padding(folder_path)

#Bancor
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    pdf_folder_path = 'bancor/bancor_assets'
    pdf_to_image(pdf_folder_path)

    extract_image_to_text(image_path)

    lines = extracted_text.strip().split('\n')

    #for proof number
    line_3 = lines[3] 
    #for payer name
    line_11 = lines[11]
    line_11 = re.sub(r'^\w+:+\s+\b', '', line_11)

    # Define regular expression patterns for different types of data to extract

    bank_pattern = 'Bancor'
    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
    proof_number_pattern = line_3
    
    payer_name_pattern = line_11
    
    cuit_pattern =  r'\b\d{11}\b'

    # Extract information using the regEx patterns
    details_regEx_patterns()

    # Extract the date, amount, and CUIT number if any are found
    extract_details()

    amount = amounts_found[1] if amounts_found else None
    
    # return extracted_details
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

# Define the folder containing the images
folder_path = 'bancor/bancor_assets'

#function 
check_image_and_padding(folder_path)

#hsbc
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    # pdf_folder_path = 'hsbc/hsbc_assets'  #uncomment later
    # pdf_to_image(pdf_folder_path)

    extract_image_to_text(image_path)

    lines = extracted_text.strip().split('\n')
    #for proof number
    # line_3 = lines[3] 
    #for payer name
    line_11 = lines[11]
    line_11 = re.sub(r'^\w+:+\s+\b', '', line_11)


    # Define regular expression patterns for different types of data to extract

    bank_pattern = 'usec' #HSBC is detected as that

    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'

    amount_pattern = r'\bARS\s*\d+\.?\d*'  # have to Replace ARS with $

    proof_number_pattern = r'\b\d{4}\b'
    
    payer_name_pattern = "None"
    
    cuit_pattern =  "None"

    # Extract information using the regEx patterns
    details_regEx_patterns()

    # Extract the date, amount, and CUIT number if any are found
    extract_details()

    bank = "hsbc" #if bank_found else None

    amount = amounts_found[0] if amounts_found else None
    amount = amount.replace('ARS', '$') #replaces ARS with $

    #this line is important to extract the proof number
    proof_number = proof_number_found[1] if proof_number_found else None
    
    # return extracted_details
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

# Define the folder containing the images
folder_path = 'hsbc/hsbc_assets'

#function 
check_image_and_padding(folder_path)

#uala
def extract_data_from_image(image_path):

    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    # pdf_folder_path = 'hsbc/hsbc_assets'  #uncomment later
    # pdf_to_image(pdf_folder_path)

    extract_image_to_text(image_path)

    lines = extracted_text.strip().split('\n')

# for  uala.png and uala1.png formats
    if lines[0] == "afr":
        global line_8, line_10
        #payer name
        line_8 = lines[9]
        line_8 = line_8.replace("Nombre remitente ", "")
        # print(line_8)
        #for proof number
        line_10 = lines[11]
        line_10 = line_10.replace("Id Op. ", "")
        # print(line_10)
    if lines[0] == "VAD Comprobante de transferencia":
        #payer name
        line_8 = lines[8]
        line_8 = line_8.replace("Nombre remitente ", "")
        # print(line_8)
        #for proof number
        line_10 = lines[10]
        line_10 = line_10.replace("Id Op. ", "")
        # print(line_10)


    # Define regular expression patterns for different types of data to extract

    bank_pattern = 'uala' #HSBC is detected as that

    date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'

    amount_pattern = r'\$\s*\d+\.?\d*'  # have to Replace ARS with $

    proof_number_pattern = line_10
    
    payer_name_pattern = line_8
    
    cuit_pattern =  "None" #MARTIN SAID Cuit IS FILLED MANUALLY 

    # Extract information using the regEx patterns
    details_regEx_patterns()

    # Extract the date, amount, and CUIT number if any are found
    extract_details()

    amount = amounts_found[0] if amounts_found else None
    amount = amount.replace("$", "$ ") #replaces '$' with '$ '
    
    # return extracted_details
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

# Define the folder containing the images
folder_path = 'uala/uala_assets'

#function 
check_image_and_padding(folder_path)


# sum and adding and saving and opening the file
print(f"total sum: {last_amt}")
print(f"Current total sum: {sum(last_amt)}")

total_sum = sum(last_amt)
# Append the total sum row to the Excel sheet
total_sum_row = ['TOTAL', None, f'${total_sum}', None, None, None, None]
ws.append(total_sum_row)

from datetime import date

today = date.today()
formatted_date = today.strftime("%d_%m_%Y")

# Specify the path to the extracted file
extracted_file_path = f'{formatted_date}_extracted_info.xlsx'


# Save the Excel file
wb.save(extracted_file_path)
import subprocess
import os

# Open the file with LibreOffice in linux
subprocess.run(['libreoffice', extracted_file_path])

# Open the file with Microsoft Excel in windows
#os.system("start EXCEL.EXE extracted_info.xlsx")
