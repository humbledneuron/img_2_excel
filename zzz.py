import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from decimal import Decimal
import cv2
import pytesseract
from pdf2image import convert_from_path
import re

# global variable for serial number
index = 1
last_amt = []  # list

folder_path = 'zource'
image_path = 'zource'

all_files = [file for file in os.listdir(folder_path)]
print(all_files)

# Get the list of PDF files in the folder and convert them to images
def pdf_to_image(pdf_folder_path):
    pdf_files = [file for file in os.listdir(pdf_folder_path) if file.endswith('.pdf')]

    # Convert each PDF file to images
    for pdf_file in pdf_files:
        # Construct the file paths
        pdf_file_path = os.path.join(pdf_folder_path, pdf_file)
        image_file_path = os.path.join(pdf_folder_path, os.path.splitext(pdf_file)[0] + '.png')

        # Convert PDF to list of PIL images
        images = convert_from_path(pdf_file_path)

        # Save each page of the PDF as an image file
        for i, image in enumerate(images):
            image.save(image_file_path, 'PNG')

    print('PDFs converted to images successfully.')

# Extract the image to text
def extract_image_to_text(image_path):
    global extracted_text

    image = cv2.imread(image_path)
    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    extracted_text = pytesseract.image_to_string(gray_image)

# Return the extracted details
def get_extracted_details(bank, date, amount, payer, cuit, proof_number):
    return {
        'BANCO': bank,
        'FECHA': date,
        'IMPORTE': amount,
        'TITULAR': payer,
        'CUIT': cuit,
        'NRO COMPROBANTE': proof_number
    }

# New Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Extracted Data"

# Add headers to the Excel sheet
headers = ['Serie', 'FECHA', 'IMPORTE', 'NRO COMPROBANTE', 'TITULAR', 'CUIT', 'BANCO']
ws.append(headers)

# Apply bold style to the header cells
bold_font = Font(bold=True)
for cell in ws[1]:
    cell.font = bold_font

# Apply padding to all cells
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Check if image exists and adjust the padding
def check_image_and_padding(folder_path):
    global index, last_amt

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for filename in os.listdir(folder_path):
        if filename.endswith(".jpg") or filename.endswith(".png") or filename.endswith(".jpeg"):
            # Extract data from the image
            image_path = os.path.join(folder_path, filename)
            extracted_data = extract_data_from_image(image_path)

            row_data = [index]  # Add the serial number as the first element
            row_data.extend([extracted_data.get(header) for header in headers[1:]])
            ws.append(row_data)
            index += 1

            # Update total sum formula dynamically
            if extracted_data['IMPORTE']:
                amt = re.sub(r'[^\$\s0-9.]', '', extracted_data['IMPORTE'])
                num_amt = Decimal(re.sub(r'[^\d.]', '', amt))
                last_amt.append(num_amt)
                print(f"Current total sum: {sum(last_amt)}")

            # Apply yellow fill to cells that are empty in the row
            if any(is_empty(cell.value) for cell in ws[ws.max_row]):
                for cell in ws[ws.max_row]:
                    if is_empty(cell.value):
                        cell.fill = yellow_fill

    # Automatically adjust column widths based on content
    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Adjust the multiplier as needed
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

# Function to check if a cell is empty
def is_empty(cell_value):
    return cell_value is None or cell_value == ''

# Process each image in the folder
def extract_data_from_image(image_path):
    extract_image_to_text(image_path)

    lines = extracted_text.split('\n')
    line = lines[0].lower()

    if line == 'bancopatagonia':
        bank_pattern = 'Bancopatagonia'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{10}\b'
        payer_name_pattern = lines[11]
        cuit_pattern = 'None'  # MARTIN SAID Cuit IS FILLED MANUALLY

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        bank = bank_found[0] if bank_found else None
        date = dates_found[0] if dates_found else None
        amount = amounts_found[1] if amounts_found else None
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

        return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

    if line == 'galicia':
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{11}\b'
        payer_name_pattern = lines[8]
        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

        bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
        dates_found = re.findall(date_pattern, extracted_text)
        amounts_found = re.findall(amount_pattern, extracted_text)
        payer_name_found = re.findall(payer_name_pattern, extracted_text)
        cuit_found = re.findall(cuit_pattern, extracted_text)
        proof_number_found = re.findall(proof_number_pattern, extracted_text)

        bank = bank_found[0] if bank_found else None
        date = dates_found[0] if dates_found else None
        amount = amounts_found[0] if amounts_found else None
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None

        if any(len(p) == 9 for p in proof_number_found) and any(len(p) == 11 for p in proof_number_found):
            proof_number = next((p for p in proof_number_found if len(p) == 9), None)
        else:
            proof_number = proof_number_found[0] if proof_number_found else None

        return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

# Process each image in the folder
check_image_and_padding(folder_path)
wb.save('Extracted_Data.xlsx')
