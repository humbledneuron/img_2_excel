import os, shutil, time, msvcrt
from openpyxl import Workbook
from pdf2image import convert_from_path
from openpyxl.styles import Font, Alignment, PatternFill
import cv2
import pytesseract
import re
from decimal import Decimal
#from tkinter import Tk, Label, Button
#from tkinter.filedialog import askdirectory
import tkinter  as tk

# Global variables
index = 1
last_amt = []
folder_path = 'tickets del der'
undetected_folder = 'undetected'
processed_files = set()

#sends the non detected images to the undetected folder in the source
undetected_folder = os.path.join(folder_path, undetected_folder)

#developer contact address


print(' ')
print("for any issues or future bulk bank extraction updates, please reach us out here...")
print("Bhanuprakash -- blogger.thomasjerry@gmail.com (main)")
print("Gaganraj -- gaganraj2002@gmail.com")

print(' ')
print("please wait.. the software is doing it's work")
print(' ')

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Tickets Extracted Data"

# Add headers to Excel sheet
headers = ['Serie', 'FECHA', 'IMPORTE', 'NRO COMPROBANTE', 'TITULAR', 'CUIT', 'BANCO']
ws.append(headers)

# Apply styles to header cells
bold_font = Font(bold=True)
for cell in ws[1]:
    cell.font = bold_font

# Apply padding to cells
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

#using the pdf2image library to convert pdf to image
# Get the list of PDF files in the folder
def pdf_to_image(pdf_folder_path):    

    pdf_files = [file for file in os.listdir(pdf_folder_path) if file.endswith('.pdf')]
    
    #time.sleep(1)
    
#    print("converting the PDFs into images...")
#    print(' ')


    # Convert each PDF file to images
    for pdf_file in pdf_files:
        # Construct the file paths
        pdf_file_path = os.path.join(pdf_folder_path, pdf_file)
        
        # Convert PDF to list of PIL images
        images = convert_from_path(pdf_file_path, first_page=1, last_page=1)
    
        image_file_path = os.path.join(pdf_folder_path, f"{os.path.splitext(pdf_file)[0]}_page1.png")
        images[0].save(image_file_path, 'PNG')

        #print(f'Saved {image_file_path}')

# Function to extract text from an image using Tesseract OCR
def extract_image_to_text(image_path):
    image = cv2.imread(image_path)
    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    extracted_text = pytesseract.image_to_string(gray_image)
    return extracted_text

#regEx patterns
def details_regEx_patterns():
    global  dated_pattern, amount_pattern, payer_name_pattern, cuit_pattern, proof_number_pattern, extracted_text, bank_found, dateds_found, amounts_found, payer_name_found, cuit_found, proof_number_found
    dateds_found = re.findall(dated_pattern, extracted_text)
    amounts_found = re.findall(amount_pattern, extracted_text)
    payer_name_found = re.findall(payer_name_pattern , extracted_text)
    cuit_found = re.findall(cuit_pattern, extracted_text)
    proof_number_found = re.findall(proof_number_pattern, extracted_text)
    
    return  dateds_found, amounts_found, payer_name_found, cuit_found, proof_number_found

# Function to move undetected images to a separate folder
def move_to_undetected(image_path, folder_path):
    # Create the "undetected" folder if it doesn't exist
    undetected_folder = os.path.join(folder_path, 'undetected') #"undetected"
    os.makedirs(undetected_folder, exist_ok=True)

    # Get the file name from the image path
    file_name = os.path.basename(image_path)

    # Move the image to the "undetected" folder
    shutil.move(image_path, os.path.join(undetected_folder, file_name))

# Function to extract data from an image based on its content
def extract_data_from_image(image_path):
    global index, extracted_text, processed_files
    extracted_text = extract_image_to_text(image_path)
    lines = extracted_text.split('\n')

    bank_name = None
    dated = None
    amount = None
    payer = None
    cuit = None
    proof_number = None

    for line in lines:
        if "bancopatagonia" in line.lower():
            bank_name = "Bancopatagonia"
            #print(image_path)
            break
        elif "galicia" in line.lower():
            bank_name = "Galicia"
            #print(image_path)
            break
        elif "mercado pago" in line.lower():
            bank_name = "Mercado pago"
            #print(image_path)
            break
        
         #santander is in many banks, thats why this
        elif "�> santander" in line.lower():
            bank_name = "Santander"
            #print(image_path)
            break
        elif "«> santander" in line.lower():
            bank_name = "Santander"
            #print(image_path)
            break
        elif "�>" in line.lower():
            bank_name = "Santander"
            #print(image_path)
            break                        
        elif "® santander" in line.lower(): #santander is in many banks, thats why this
            bank_name = "Santander"
            #print(image_path)
            break
        elif "AS x" in line.lower():
            bank_name = "Santander"
            #print(image_path)
            break   
        elif "> santander" in line.lower():
            bank_name = "Santander"
            #print(image_path)
            break

        #because it is not dectecting the CUenta DNI in same line
        elif '"4 dni' in line.lower():
            bank_name = "Cuenta DNI"
            #print(image_path)
            break
        elif "v4 dni" in line.lower():         
            bank_name = "Cuenta DNI"
            print(image_path)
            break
        elif "4 dni" in line.lower():         
            bank_name = "Cuenta DNI"
            #print(image_path)
            break


        # hold this becauze sometimes it has the 'Santander' word in the middle of the banks name
        elif "bna" in line.lower():
            bank_name = "BNA"
            #print(image_path)
            break

        # hold this becauze sometimes it has the 'Santander' word in the middle of the banks name
        elif "supervielle" in line.lower():
            bank_name = "SUPERVIELLE"
            #print(image_path)
            break

        elif "bancociudad" in line.lower():
            bank_name = "BancoCiudad"
            #print(image_path)
            break

        elif "banco santa fe" in line.lower():
            bank_name = "Banco Santa Fe"
            print(image_path)
            break

        elif "bbva" in line.lower():
            bank_name = "BBVA"
            #print(image_path)
            break

        elif "naranja x" in line.lower():
            bank_name = "Naranja X"
            #print(image_path)
            break

        elif "banco credicoop coop. ltdo" in line.lower():
            bank_name = "Banco Credicoop Coop. Ltdo"
            #print(image_path)
            break

        elif "personal pay" in line.lower(): #€
            lines[0] == 'personal pay'
            bank_name = "Personal Pay"
            #print(image_path)
            break

        elif "bancor" in line.lower():
            bank_name = "Bancor"
            #print(image_path)
            break

            #hsbc has two patterns, thats why this
        elif "xp" in line.lower(): # or 'xp uss' 'ars' coz #HSBC is not detected correctly either "xp" as bank symbol and "ARS" as currency
            bank_name = "HSBC"
            #print(image_path)
            break
        elif "xp usec" in line.lower(): # or 'xp uss' 'ars' coz #HSBC is not detected correctly either "xp" as bank symbol and "ARS" as currency
            bank_name = "HSBC"
            #print(image_path)
            break        

        elif "uala" in line.lower(): 
            bank_name = "Uala" 
            #print(image_path)
            break
        
        #some times for uala, U is v in detection 
        elif "vala" in line.lower(): 
            bank_name = "Uala" 
            #print(image_path)
            break

        elif "macro" in line.lower(): 
            bank_name = "Macro" 
            #print(image_path)
            break


    if bank_name == "Bancopatagonia":
        global bank,  dated_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern, bank_found, dateds_found, amounts_found, payer_name_found, cuit_found, proof_number_found
        #dated, amount, payer, cuit, proof_number #bank_pattern,
        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{10}\b'
        payer_name_pattern = lines[11]
        cuit_pattern = 'None'  # MARTIN SAID Cuit IS FILLED MANUALLY

        # Extract information using the regEx patterns
        details_regEx_patterns()
    
        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[1] if amounts_found else None
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "Galicia":
        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d[\d,\.,\d,\.]*' #r' \$\s*\d+\.\d{2}' #r'\$\s*\d+\.?\d*+'
        proof_number_pattern = r'\b\d{9,11}\b'
        payer_name_pattern = lines[9] #lines[8] #maybe for the other pattern the payer_name 
        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

        # Extract information using the regEx patterns
        details_regEx_patterns()

        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[1] if cuit_found else None

            # Check if both 9-digit and 11-digit proof numbers are present
        if any(len(p) == 9 for p in proof_number_found) and any(len(p) == 11 for p in proof_number_found):
            # Extract the 9-digit proof number if both are present
            proof_number = next((p for p in proof_number_found if len(p) == 9), None)
        else:
            # Extract the first found proof number (either 9-digit or 11-digit)
            proof_number = proof_number_found[0] if proof_number_found else None

#one error
    elif bank_name == "Mercado pago":
        dated_pattern = r'\b\d{1,2} de [a-z]+ \d{4}\b'
        amount_pattern = r'[\$¢]\s*\d[\d,\.]+' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{11}\b'
        payer_name_pattern = r'(?:de )?([A-Z][a-z]+(?: [A-Z][a-z]+)*)'
        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

        # Extract information using the regEx patterns
        details_regEx_patterns()

        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('¢', '$')
        amount = amount.replace('.', '')
        amount = amount.split(',')[0]
        payer = payer_name_found[1] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

   # hold this becauze it has the 'Santander' word in the middle of the text
    elif bank_name == "BNA":
        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = 'None' #MARTIN SAID NAME IS FILLED MANUALLY 
        cuit_pattern = r'\b\d{11}\b' #MARTIN SAID Cuit IS FILLED MANUALLY 

        # Extract information using the regEx patterns
        details_regEx_patterns()

        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None 

    # hold this becauze it has the 'Santander' word in the middle of the text
    elif bank_name == "SUPERVIELLE":
        dated_pattern = r'\b\d{1,2}\s+[A-Za-z]+\s+\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{4}\b'
        payer_name_pattern = 'None' #MARTIN SAID NAME IS FILLED MANUALLY 
        cuit_pattern = 'None' #MARTIN SAID Cuit IS FILLED MANUALLY 

        # Extract information using the regEx patterns
        details_regEx_patterns()

        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "BancoCiudad":

        lines_33 = lines[34] + lines[35]

        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = lines_33
        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

        # Extract information using the regEx patterns
        details_regEx_patterns()

        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('.', '')
        payer = re.sub(r'[0-9,-]+', '', payer_name_pattern)
        cuit = cuit_found[1] if cuit_found else None
        proof_number = proof_number_found[1] if proof_number_found else None

    elif bank_name == "Banco Santa Fe":

        line_24 = lines[24] 

        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = line_24
        cuit_pattern = r'\b\d{11}\b'

        # Extract information using the regEx patterns
        details_regEx_patterns()

        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "BBVA":

        if lines[3] == 'Transferiste a':
            #for payer name
            line_4 = lines[13]
            line_4 = re.sub(r'^\w+\s+\b', '', line_4)

            #for proof number
            line_3 = lines[9]
            line_3 = re.sub(r'^\w+\s+\w+\s+\w+\s', '', line_3) #Numero de referencia 

            #for amount
            line_12 = lines[7]
            # line_12 = re.sub(r'\b,00+$', '', line_12)
            line_12 = line_12.split(',')[0]

            #bank_pattern = 'BBVA'
            dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
            amount_pattern = line_12
            payer_name_pattern = line_4
            proof_number_pattern = line_3

            cuit_pattern = 'None'

        else:
            #for payer name
            line_4 = lines[4]
            line_4 = re.sub(r'^\w+:+\s+\b', '', line_4)#1st word
            line_4 = re.sub(r'\b\s+\w+$', '', line_4)#last word

            #for proof number
            line_3 = lines[3]
            Cuenta_Origen = r'^Cuenta Origen:\s+CC\s+\$\s+\d{4}-\d{6}\/\d{1}\s+'
            line_3 = re.sub(Cuenta_Origen, '', line_3)
            line_3 = re.sub(r' ' , '', line_3)

            #for amount
            line_12 = lines[12]
            #removes 1st word
            line_12 = re.sub(r'^\w+:+\s+\b', '', line_12)
            #adds $ 
            line_12 = '$ ' + line_12
            line_12 = re.sub(r'\b,00+$', '', line_12)


            #bank_pattern = 'BBVA'
            dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
            #something is wrong here
            amount_pattern = line_12
            payer_name_pattern = line_4
            proof_number_pattern = line_3

            cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

        # Extract information using the regEx patterns
        details_regEx_patterns()

        dated = dateds_found[0] if dateds_found else None
        amount = line_12
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = line_3

    elif bank_name == "Naranja X":

        if lines[0] == 'fod': #or lines[0] == 'Foy':
            #for extracting payer name
            line_14 = lines[13]

        elif lines[0] == 'Foy':
            #for extracting payer name
            line_14 = lines[14]

        elif lines[0] == '<':
            #for extracting payer name
            line_14 = lines[12]

        dated_pattern = r'\b\d{1,2}\/+[A-Za-z]+\/+\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' 
        payer_name_pattern = line_14
        proof_number_pattern = r'\b\d{10}\b'

        cuit_pattern = r'\b\d{11}\b'

        # Extract information using the regEx patterns
        details_regEx_patterns()

        dated = dateds_found[0] if dateds_found else None#line_8 
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('.', '')
        payer = line_14 #payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "Banco Credicoop Coop. Ltdo":

        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        payer_name_pattern = 'None'
        proof_number_pattern = r'\b\d{9}\b'

        cuit_pattern = 'None'

        # Extract information using the regEx patterns
        details_regEx_patterns()

        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[1] if amounts_found else None
        amount = amount.replace('$', '$ ')
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "Personal Pay":

        #if the amount has 1 in it, the OCR is dectecting it as ] that's why we have to check the first line
        #if lines[0] == '� personal pay':

            #for amount
            line_4 = lines[4]
            line_4 = re.sub(r'"', '', line_4) # remove '' in the end of the amount
            line_4 = re.sub(r'[A-Za-z]', '', line_4)
            #line_4 = re.sub(r'(?<=\$)(?=\d)', r' ', line_4) # adds '$ ' in place of '$'
            line_4 = '$ ' + line_4

            line_39 = lines[39] #43 for first proof number
            line_40 = line_39 + lines[40] #44 for second proof number

            #for payer name
            line_28 = lines[28]


            dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
            amount_pattern = line_4
            payer_name_pattern = line_28
            proof_number_pattern = line_40
            cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

            # Extract information using the regEx patterns
            details_regEx_patterns()

            dated = dateds_found[0] if dateds_found else None
            amount = line_4
            amount = amount.replace('.', '')
            payer = line_28
            cuit = cuit_found[0] if cuit_found else None
            proof_number = line_40
       
    elif bank_name == "Bancor":

        #for proof number
        line_3 = lines[3] 
        #for payer name
        line_11 = lines[11]
        line_11 = re.sub(r'^\w+:+\s+\b', '', line_11)

        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = line_3
        payer_name_pattern = line_11
        cuit_pattern = r'\b\d{11}\b'  # MARTIN SAID Cuit IS FILLED MANUALLY

        # Extract information using the regEx patterns
        details_regEx_patterns()

        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[1] if amounts_found else None
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "HSBC":

        if lines[0] == 'Xp usec':
            #for payer name
            line_11 = lines[4]
            # line_11 = re.sub(r'^\w+\s+:+\s+\b', '', line_11)
            line_11 = re.sub(r'^.*?\s*:\s*', '', line_11)
            # line_11 = re.sub(r'Razén Social: ', '', line_11)

            dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
            amount_pattern =  r'\bARS\s*\d+\.?\d*'  # have to Replace ARS with $
            proof_number_pattern = r'\b\d{8}\b'
            payer_name_pattern = line_11
            cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

            # Extract information using the regEx patterns
            details_regEx_patterns()

            dated = dateds_found[0] if dateds_found else None
            amount = amounts_found[1] if amounts_found else None
            amount = amount.replace('ARS', '$ ') #replaces 'ARS' with '$ '
            amount = amount.split('.')[0] #amount.replace('.', '')
            payer = payer_name_found[0] if payer_name_found else None
            cuit = cuit_found[0] if cuit_found else None
            proof_number = proof_number_found[2] if proof_number_found else None        
        
        else:

            dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
            amount_pattern =  r'\bARS\s*\d+\.?\d*'  # have to Replace ARS with $
            proof_number_pattern = r'\b\d{4}\b'
            payer_name_pattern = "None"
            cuit_pattern = "None"  # MARTIN SAID Cuit IS FILLED MANUALLY

            # Extract information using the regEx patterns
            details_regEx_patterns()

            dated = dateds_found[0] if dateds_found else None
            amount = amounts_found[0] if amounts_found else None
            amount = amount.replace('ARS', '$ ') #replaces '$' with '$ '
            amount = amount.replace('.', '')
            amount = amount.split(',')[0] #cuts off upto ,
            payer = payer_name_found[0] if payer_name_found else None
            cuit = cuit_found[0] if cuit_found else None
            proof_number = proof_number_found[1] if proof_number_found else None

    elif bank_name == "Uala":

        # for  uala.png and uala1.png formats
        if lines[0] == "afr":
            # global line_8, line_10
            #payer name
            line_8 = lines[9]
            line_8 = line_8.replace("Nombre remitente ", "")
            # print(line_8)
            #for proof number
            line_10 = lines[11]
            line_10 = line_10.replace("Id Op. ", "")
            # print(line_10)

            dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
            amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
            proof_number_pattern = line_10
            payer_name_pattern = line_8
            cuit_pattern = "None"  # MARTIN SAID Cuit IS FILLED MANUALLY

            # Extract information using the regEx patterns
            details_regEx_patterns()

            dated = dateds_found[0] if dateds_found else None
            amount = amounts_found[0] if amounts_found else None
            amount = amount.replace('$', '$ ') #replaces '$' with '$ '
            amount = amount.replace('.', '')
            payer = payer_name_found[0] if payer_name_found else None
            cuit = cuit_found[0] if cuit_found else None
            proof_number = proof_number_found[0] if proof_number_found else None

        # if lines[0] == "VAD Comprobante de transferencia":
        else:
            #payer name
            line_8 = lines[8]
            line_8 = line_8.replace("Nombre remitente ", "")
            # print(line_8)
            #for proof number
            line_10 = lines[10]
            line_10 = line_10.replace("Id Op. ", "")

            dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
            amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
            proof_number_pattern = line_10
            payer_name_pattern = line_8
            cuit_pattern = "None"  # MARTIN SAID Cuit IS FILLED MANUALLY

            # Extract information using the regEx patterns
            details_regEx_patterns()

            dated = dateds_found[0] if dateds_found else None
            amount = amounts_found[0] if amounts_found else None
            amount = amount.replace('$', '$ ') #replaces '$' with '$ '
            amount = amount.replace('.', '')
            payer = payer_name_found[0] if payer_name_found else None
            cuit = cuit_found[0] if cuit_found else None
            proof_number = proof_number_found[0] if proof_number_found else None

        #one error, language not supported 
    elif bank_name == "Santander":
        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = 'None' #MARTIN SAID NAME IS FILLED MANUALLY 
        cuit_pattern = 'None' #MARTIN SAID Cuit IS FILLED MANUALLY 

        # Extract information using the regEx patterns
        details_regEx_patterns()

        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "Cuenta DNI":

        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        cuit_pattern = r'\b\d{11}\b'

        # if lines[0] == "fo al":
        # if lines[1] == "v4 DNI":
        #payer name
        lines_10 = lines[10]
            
        proof_number_pattern = r'\b\d{6,8}\b' #6,8
        payer_name_pattern = lines_10  

        # Extract information using the regEx patterns
        details_regEx_patterns()

        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "Macro":

        #for payer name
        line_14 = lines[14]
        
        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern =  r'\$\s*\d+\,?\d*' 
        #r'\$\s*\d[\d,]*' #r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{9}\b'
        payer_name_pattern = line_14
        cuit_pattern = "None" #because it's not detecting correctly #r"\b\d{11}\b"

        # Extract information using the regEx patterns
        details_regEx_patterns()
        
        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[1] if amounts_found else None
        amount = re.sub(',', '', amount) #amount.replace(',', '')
        amount = amount.split('.')[0]
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    else:
        move_to_undetected(image_path, folder_path)
    
    # Return the extracted data
    return {
        'BANCO': bank_name,
        'FECHA': dated,
        'IMPORTE': amount,
        'NRO COMPROBANTE': proof_number,
        'TITULAR': payer,
        'CUIT': cuit
    }

# Function to check if a cell is empty
def is_empty(cell_value):
    return cell_value is None or cell_value == ''

# Function to check images in the folder and add data to Excel sheet
def check_image_and_padding(folder_path):
    global index, last_amt
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for filename in os.listdir(folder_path):
        if filename.lower().endswith(('.jpg', '.jpeg', '.png')) and filename not in processed_files:
            processed_files.add(filename)
            # Extract data from the image
            image_path = os.path.join(folder_path, filename)
            extracted_data = extract_data_from_image(image_path)
            if extracted_data:
                # Add data to Excel sheet
                row_data = [index]
                row_data.extend([extracted_data.get(header, '') for header in headers[1:]])
                ws.append(row_data)
                index += 1
                            # Updated total sum formula dynamically
                if extracted_data['IMPORTE']:
                    amt = re.sub(r'[^\$\s0-9.]', '', extracted_data['IMPORTE'])
                    amt = re.sub(r'[^\d.,]', '', amt) #r'[^\d.]'
                    num_amt = Decimal(amt)
                    # total_sum_formula += num_amt

                    last_amt.append(num_amt)
                    #print(f"Current total sum: {sum(last_amt)}")

                # Apply yellow fill to empty cells
                if any(is_empty(cell.value) for cell in ws[ws.max_row]):
                    for cell in ws[ws.max_row]:
                        if is_empty(cell.value):
                            cell.fill = yellow_fill
    # Adjust column widths
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

# Process images in the folder and save Excel file
pdf_to_image(folder_path)

#print(' ')
#print("Pdfs are converted to images...")

check_image_and_padding(folder_path)



# sum and adding and saving and opening the file
one_space_row = [None, None, None, None, None, None, None]
ws.append(one_space_row)


total_sum = sum(last_amt)

# Create a bold font style
bold_font = Font(bold=True)

# Append the total sum row to the Excel sheet
total_sum_row = ['SUMA TOTAL', None, f'${total_sum}', None, None, None, None]
ws.append(total_sum_row)

#extracts todays date and makes a folder out of it    
from datetime import date

today = date.today()
formatted_dated = today.strftime("%d_%m_%Y")

# Specify the path to the extracted file
extracted_file_path = os.path.join(folder_path, 'detected')
os.makedirs(extracted_file_path, exist_ok=True)
extracted_file_path = os.path.join(extracted_file_path, f'{formatted_dated}_extracted_info.xlsx')


# Save the Excel file
wb.save(extracted_file_path)

print(" ")
print(" ")
print(f"the file saved at \n {extracted_file_path}, \n please re-check, the excel sheet, because there might be some errors, and you said some details should be filled manually ")

def is_enter_pressed():
    return msvcrt.kbhit() and ord(msvcrt.getch()) == 13


# Wait for the user to press Enter to exit
print(" ")
print("Press Enter to exit...")


# Keep blinking the last dot until Enter is pressed
while True:
    print('.', end='', flush=True)
    time.sleep(0.5)  # Adjust blinking speed as needed
    
    if is_enter_pressed():
        print("\nExiting program...")
        break


# Open the file with Microsoft Excel in windows
#os.system(f'start "{extracted_file_path}"')

