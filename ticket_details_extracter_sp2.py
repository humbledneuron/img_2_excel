import os, shutil, time
from openpyxl import Workbook
from pdf2image import convert_from_path
from openpyxl.styles import Font, Alignment, PatternFill
import cv2
import pytesseract
import re
from decimal import Decimal
from tkinter import Tk, Label, Button
from tkinter.filedialog import askdirectory

# Global variables
index = 1
last_amt = []
folder_path = ''
undetected_folder = 'undetected'
processed_files = set()

#opens a window to select the folder
# Create a Tkinter root window
root = Tk()
# Hide the root window
root.withdraw()
# Initialize
folder_path = askdirectory()
# Print the selected folder path
print(' ')#this prints nothing
print("Selected folder:", folder_path)
print(' ')#this prints nothing

#sends the non detected images to the undetected folder in the source
undetected_folder = os.path.join(folder_path, undetected_folder)

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Tickets Extracted Data"

# Add headers to Excel sheet
headers = ['Serie', 'FECHA', 'IMPORTE', 'NRO COMPROBANTE', 'TITULAR', 'CUIT', 'BANCO']
ws.append(headers)

# Apply styles to header cells
bold_font = Font(bold=True)
for cell in ws[1]:
    cell.font = bold_font

# Apply padding to cells
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

#using the pdf2image library to convert pdf to image
# Get the list of PDF files in the folder
def pdf_to_image(pdf_folder_path):    

    pdf_files = [file for file in os.listdir(pdf_folder_path) if file.endswith('.pdf')]
    
    time.sleep(1)
    print("converting the PDFs into images...")
    print(' ')

    # Convert each PDF file to images
    for pdf_file in pdf_files:
        # Construct the file paths
        pdf_file_path = os.path.join(pdf_folder_path, pdf_file)
        
        # Convert PDF to list of PIL images
        images = convert_from_path(pdf_file_path, first_page=1, last_page=1)
    
        image_file_path = os.path.join(pdf_folder_path, f"{os.path.splitext(pdf_file)[0]}_page1.png")
        images[0].save(image_file_path, 'PNG')

        print(f'Saved {image_file_path}')


#####Emergency only ##### specify the path of the tesseract
#pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'


# Function to extract text from an image using Tesseract OCR
def extract_image_to_text(image_path):
    image = cv2.imread(image_path)
    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    extracted_text = pytesseract.image_to_string(gray_image)
    return extracted_text

#regEx patterns
def details_regEx_patterns():
    global  dated_pattern, amount_pattern, payer_name_pattern, cuit_pattern, proof_number_pattern, extracted_text, bank_found, dateds_found, amounts_found, payer_name_found, cuit_found, proof_number_found
    #bank_pattern,
    #bank_found = re.findall(#bank_pattern, extracted_text, re.IGNORECASE)
    dateds_found = re.findall(dated_pattern, extracted_text)
    amounts_found = re.findall(amount_pattern, extracted_text)
    payer_name_found = re.findall(payer_name_pattern , extracted_text)
    cuit_found = re.findall(cuit_pattern, extracted_text)
    proof_number_found = re.findall(proof_number_pattern, extracted_text)
    
    # bank_found,
    return  dateds_found, amounts_found, payer_name_found, cuit_found, proof_number_found

# Function to move undetected images to a separate folder
def move_to_undetected(image_path, folder_path):
    # Create the "undetected" folder if it doesn't exist
    undetected_folder = os.path.join(folder_path, 'undetected') #"undetected"
    os.makedirs(undetected_folder, exist_ok=True)

    # Get the file name from the image path
    file_name = os.path.basename(image_path)

    # Move the image to the "undetected" folder
    shutil.move(image_path, os.path.join(undetected_folder, file_name))

# Function to extract data from an image based on its content
def extract_data_from_image(image_path):
    global index, extracted_text, processed_files
    extracted_text = extract_image_to_text(image_path)
    lines = extracted_text.split('\n')

    bank_name = None
    dated = None
    amount = None
    payer = None
    cuit = None
    proof_number = None

    for line in lines:
        if "bancopatagonia" in line.lower():
            print(line.lower())
            bank_name = "Bancopatagonia"
            print(bank_name)
            print(image_path)
            break
        elif "galicia" in line.lower():
            print(line.lower())
            bank_name = "Galicia"
            print(bank_name)
            print(image_path)
            break
        elif "mercado pago" in line.lower():
            print(line.lower())
            bank_name = "Mercado pago"
            print(bank_name)
            print(image_path)
            break
        
         #santander is in many banks, thats why this
        elif "�> santander" in line.lower():
            print(line.lower())
            bank_name = "Santander"
            print(bank_name)
            print(image_path)
            break
        elif "«> santander" in line.lower():
            print(line.lower())
            bank_name = "Santander"
            print(bank_name)
            print(image_path)
            break
        elif "�>" in line.lower():
            print(line.lower())
            bank_name = "Santander"
            print(bank_name)
            print(image_path)
            break                        
        elif "® santander" in line.lower(): #santander is in many banks, thats why this
            print(line.lower())
            bank_name = "Santander"
            print(bank_name)
            print(image_path)
            break
        elif "AS x" in line.lower():
            print(line.lower())
            bank_name = "Santander"
            print(bank_name)
            print(image_path)
            break   
        elif "> santander" in line.lower():
            print(line.lower())
            bank_name = "Santander"
            print(bank_name)
            print(image_path)
            break

        #because it is not dectecting the CUenta DNI in same line
        elif '"4 dni' in line.lower():
            print(line.lower())
            bank_name = "Cuenta DNI"
            print(bank_name)
            print(image_path)
            break
        elif "v4 dni" in line.lower():         
            print(line.lower())
            bank_name = "Cuenta DNI"
            print(bank_name)
            print(image_path)
            break
        elif "4 dni" in line.lower():         
            print(line.lower())
            bank_name = "Cuenta DNI"
            print(bank_name)
            print(image_path)
            break


        # hold this becauze sometimes it has the 'Santander' word in the middle of the banks name
        elif "bna" in line.lower():
            print(line.lower())
            bank_name = "BNA"
            print(bank_name)
            print(image_path)
            break

        # hold this becauze sometimes it has the 'Santander' word in the middle of the banks name
        elif "supervielle" in line.lower():
            print(line.lower())
            bank_name = "SUPERVIELLE"
            print(bank_name)
            print(image_path)
            break

        elif "bancociudad" in line.lower():
            print(line.lower())
            bank_name = "BancoCiudad"
            print(bank_name)
            print(image_path)
            break

        elif "banco santa fe" in line.lower():
            print(line.lower())
            bank_name = "Banco Santa Fe"
            print(bank_name)
            print(image_path)
            break

        elif "bbva" in line.lower():
            print(line.lower())
            bank_name = "BBVA"
            print(bank_name)
            print(image_path)
            break

        elif "naranja x" in line.lower():
            print(line.lower())
            bank_name = "Naranja X"
            print(bank_name)
            print(image_path)
            break

        elif "banco credicoop coop. ltdo" in line.lower():
            print(line.lower())
            bank_name = "Banco Credicoop Coop. Ltdo"
            print(bank_name)
            print(image_path)
            break

        elif "personal pay" in line.lower(): #€
            lines[0] == 'personal pay'
            print(line.lower())
            bank_name = "Personal Pay"
            print(bank_name)
            print(image_path)
            break

        elif "bancor" in line.lower():
            print(line.lower())
            bank_name = "Bancor"
            print(bank_name)
            print(image_path)
            break

            #hsbc has two patterns, thats why this
        elif "xp" in line.lower(): # or 'xp uss' 'ars' coz #HSBC is not detected correctly either "xp" as bank symbol and "ARS" as currency
            print(line.lower())
            bank_name = "HSBC"
            print(bank_name)
            print(image_path)
            break
        elif "xp usec" in line.lower(): # or 'xp uss' 'ars' coz #HSBC is not detected correctly either "xp" as bank symbol and "ARS" as currency
            print(line.lower())
            bank_name = "HSBC"
            print(bank_name)
            print(image_path)
            break        

        elif "uala" in line.lower(): 
            print(line.lower())
            bank_name = "Uala" 
            print(bank_name)
            print(image_path)
            break
        
        #some times for uala, U is v in detection 
        elif "vala" in line.lower(): 
            print(line.lower())
            bank_name = "Uala" 
            print(bank_name)
            print(image_path)
            break

        elif "macro" in line.lower(): 
            print(line.lower())
            bank_name = "Macro" 
            print(bank_name)
            print(image_path)
            break


    if bank_name == "Bancopatagonia":
        global bank,  dated_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern, bank_found, dateds_found, amounts_found, payer_name_found, cuit_found, proof_number_found
        #dated, amount, payer, cuit, proof_number #bank_pattern,
        #bank_pattern = 'Bancopatagonia'
        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{10}\b'
        payer_name_pattern = lines[11]
        cuit_pattern = 'None'  # MARTIN SAID Cuit IS FILLED MANUALLY

        # Extract information using the regEx patterns
        details_regEx_patterns()
    
        #bank = bank_found[0] if bank_found else None
        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[1] if amounts_found else None
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "Galicia":
        #bank_pattern = 'Galicia'
        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{9,11}\b'
        payer_name_pattern = lines[8]
        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

        # Extract information using the regEx patterns
        details_regEx_patterns()

        #bank = bank_found[0] if bank_found else None
        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None

            # Check if both 9-digit and 11-digit proof numbers are present
        if any(len(p) == 9 for p in proof_number_found) and any(len(p) == 11 for p in proof_number_found):
            # Extract the 9-digit proof number if both are present
            proof_number = next((p for p in proof_number_found if len(p) == 9), None)
        else:
            # Extract the first found proof number (either 9-digit or 11-digit)
            proof_number = proof_number_found[0] if proof_number_found else None

#one error
    elif bank_name == "Mercado pago":
        #bank_pattern = 'Mercado pago'
        dated_pattern = r'\b\d{1,2} de [a-z]+ \d{4}\b'
        amount_pattern = r'[\$¢]\s*\d[\d,\.]+' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{11}\b'
        payer_name_pattern = r'(?:de )?([A-Z][a-z]+(?: [A-Z][a-z]+)*)'
        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

        # Extract information using the regEx patterns
        details_regEx_patterns()

        #bank = bank_found[0] if bank_found else None
        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('¢', '$')
        amount = amount.replace('.', '')
        amount = amount.split(',')[0]
        payer = payer_name_found[1] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

   # hold this becauze it has the 'Santander' word in the middle of the text
    elif bank_name == "BNA":
        #bank_pattern = 'BNA'
        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = 'None' #MARTIN SAID NAME IS FILLED MANUALLY 
        cuit_pattern = r'\b\d{11}\b' #MARTIN SAID Cuit IS FILLED MANUALLY 

        # Extract information using the regEx patterns
        details_regEx_patterns()

        #bank = bank_found[0] if bank_found else None
        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None 

    # hold this becauze it has the 'Santander' word in the middle of the text
    elif bank_name == "SUPERVIELLE":
        #bank_pattern = 'SUPERVIELLE'
        dated_pattern = r'\b\d{1,2}\s+[A-Za-z]+\s+\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{4}\b'
        payer_name_pattern = 'None' #MARTIN SAID NAME IS FILLED MANUALLY 
        cuit_pattern = 'None' #MARTIN SAID Cuit IS FILLED MANUALLY 

        # Extract information using the regEx patterns
        details_regEx_patterns()

        #bank = bank_found[0] if bank_found else None
        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "BancoCiudad":

        lines_33 = lines[34] + lines[35]

        #bank_pattern = 'BancoCiudad'
        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = lines_33
        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

        # Extract information using the regEx patterns
        details_regEx_patterns()

        #bank = bank_found[0] if bank_found else None
        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('.', '')
        payer = re.sub(r'[0-9,-]+', '', payer_name_pattern)
        cuit = cuit_found[1] if cuit_found else None
        proof_number = proof_number_found[1] if proof_number_found else None

    elif bank_name == "Banco Santa Fe":

        line_24 = lines[24] 

        #bank_pattern = 'Banco Santa Fe'
        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = line_24
        cuit_pattern = r'\b\d{11}\b'

        # Extract information using the regEx patterns
        details_regEx_patterns()

        #bank = bank_found[0] if bank_found else None
        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "BBVA":

        if lines[3] == 'Transferiste a':
            #for payer name
            line_4 = lines[13]
            line_4 = re.sub(r'^\w+\s+\b', '', line_4)

            #for proof number
            line_3 = lines[9]
            line_3 = re.sub(r'^\w+\s+\w+\s+\w+\s', '', line_3) #Numero de referencia 

            #for amount
            line_12 = lines[7]
            # line_12 = re.sub(r'\b,00+$', '', line_12)
            line_12 = line_12.split(',')[0]

            #bank_pattern = 'BBVA'
            dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
            #something is wrong here
            amount_pattern = line_12
            payer_name_pattern = line_4
            proof_number_pattern = line_3

            cuit_pattern = 'None'

        else:
            #for payer name
            line_4 = lines[4]
            line_4 = re.sub(r'^\w+:+\s+\b', '', line_4)#1st word
            line_4 = re.sub(r'\b\s+\w+$', '', line_4)#last word

            #for proof number
            line_3 = lines[3]
            Cuenta_Origen = r'^Cuenta Origen:\s+CC\s+\$\s+\d{4}-\d{6}\/\d{1}\s+'
            line_3 = re.sub(Cuenta_Origen, '', line_3)
            line_3 = re.sub(r' ' , '', line_3)

            #for amount
            line_12 = lines[12]
            #removes 1st word
            line_12 = re.sub(r'^\w+:+\s+\b', '', line_12)
            #adds $ 
            line_12 = '$ ' + line_12
            line_12 = re.sub(r'\b,00+$', '', line_12)


            #bank_pattern = 'BBVA'
            dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
            #something is wrong here
            amount_pattern = line_12
            payer_name_pattern = line_4
            proof_number_pattern = line_3

            cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

        # Extract information using the regEx patterns
        details_regEx_patterns()

        #bank = bank_found[0] if bank_found else None
        dated = dateds_found[0] if dateds_found else None
        amount = line_12
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = line_3

    elif bank_name == "Naranja X":

        if lines[0] == 'fod': #or lines[0] == 'Foy':
            #for dated
            # line_8 = lines[7]
            # line_8 = re.sub(r'^\w+\s+\w+\s+\b|\b\s-\s\d{2}:\d{2}\s+h$', '', line_8) #|\b\s-\s\d{2}:\d{2}\s+h$
            #for extracting payer name
            line_14 = lines[13]

        elif lines[0] == 'Foy':
            #for dated
            # line_8 = lines[8]
            # line_8 = re.sub(r'^\w+\s+\w+\s+\b|\b\s-\s\d{2}:\d{2}\s+h$', '', line_8)
            #for extracting payer name
            line_14 = lines[14]

        elif lines[0] == '<':
            #for dated
            # line_8 = lines[5]
            # line_8 = re.sub(r'^\w+\s+\w+\s+\b|\b\s-\s\d{2}:\d{2}\s+h$', '', line_8) #|\b\s-\s\d{2}:\d{2}\s+h$
            #for extracting payer name
            line_14 = lines[12]

        #bank_pattern = 'Naranja X'
        # dated_pattern = line_8
        dated_pattern = r'\b\d{1,2}\/+[A-Za-z]+\/+\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' 
        payer_name_pattern = line_14
        proof_number_pattern = r'\b\d{10}\b'

        cuit_pattern = r'\b\d{11}\b'

        # Extract information using the regEx patterns
        details_regEx_patterns()

        #bank = bank_found[0] if bank_found else None
        dated = dateds_found[0] if dateds_found else None#line_8 
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('.', '')
        payer = line_14 #payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "Banco Credicoop Coop. Ltdo":

        #bank_pattern = 'Banco Credicoop Coop. Ltdo' #line_2
        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        payer_name_pattern = 'None'
        proof_number_pattern = r'\b\d{9}\b'

        cuit_pattern = 'None'

        # Extract information using the regEx patterns
        details_regEx_patterns()

        #bank = bank_found[0] if bank_found else None
        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[1] if amounts_found else None
        amount = amount.replace('$', '$ ')
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "Personal Pay":

        #if the amount has 1 in it, the OCR is dectecting it as ] that's why we have to check the first line
        #if lines[0] == '� personal pay':

            #for amount
            line_4 = lines[4]
            line_4 = re.sub(r'"', '', line_4) # remove '' in the end of the amount
            line_4 = re.sub(r'[A-Za-z]', '', line_4)
            #line_4 = re.sub(r'(?<=\$)(?=\d)', r' ', line_4) # adds '$ ' in place of '$'
            line_4 = '$ ' + line_4

            line_39 = lines[39] #43 for first proof number
            line_40 = line_39 + lines[40] #44 for second proof number

            #for payer name
            line_28 = lines[28]


            #bank_pattern = 'Personal Pay'
            dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
            amount_pattern = line_4
            payer_name_pattern = line_28
            proof_number_pattern = line_40
            cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

            # Extract information using the regEx patterns
            details_regEx_patterns()

            #bank = bank_found[0] if bank_found else None
            dated = dateds_found[0] if dateds_found else None
            amount = line_4
            amount = amount.replace('.', '')
            payer = line_28
            cuit = cuit_found[0] if cuit_found else None
            proof_number = line_40
        #else: 
            #print(f"the {image_path} can't be detected")
       
    elif bank_name == "Bancor":

        #for proof number
        line_3 = lines[3] 
        #for payer name
        line_11 = lines[11]
        line_11 = re.sub(r'^\w+:+\s+\b', '', line_11)

        #bank_pattern = 'Bancor'
        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = line_3
        payer_name_pattern = line_11
        cuit_pattern = r'\b\d{11}\b'  # MARTIN SAID Cuit IS FILLED MANUALLY

        # Extract information using the regEx patterns
        details_regEx_patterns()

        #bank = bank_found[0] if bank_found else None
        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[1] if amounts_found else None
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "HSBC":

        if lines[0] == 'Xp usec':
            #for payer name
            line_11 = lines[4]
            # line_11 = re.sub(r'^\w+\s+:+\s+\b', '', line_11)
            line_11 = re.sub(r'^.*?\s*:\s*', '', line_11)
            # line_11 = re.sub(r'Razén Social: ', '', line_11)

            #bank_pattern = 'HSBC' #'usec' #HSBC is detected as that
            dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
            amount_pattern =  r'\bARS\s*\d+\.?\d*'  # have to Replace ARS with $
            proof_number_pattern = r'\b\d{8}\b'
            payer_name_pattern = line_11
            cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

            # Extract information using the regEx patterns
            details_regEx_patterns()

            # bank = "HSBC"
            dated = dateds_found[0] if dateds_found else None
            amount = amounts_found[1] if amounts_found else None
            amount = amount.replace('ARS', '$ ') #replaces '$' with '$ '
            amount = amount.split('.')[0] #amount.replace('.', '')
            payer = payer_name_found[0] if payer_name_found else None
            cuit = cuit_found[0] if cuit_found else None
            proof_number = proof_number_found[2] if proof_number_found else None        
        
        else:
            #for payer name
            # line_11 = lines[11]
            # line_11 = re.sub(r'^\w+:+\s+\b', '', line_11)

            #bank_pattern = 'HSBC' #'usec' #HSBC is detected as that
            dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
            amount_pattern =  r'\bARS\s*\d+\.?\d*'  # have to Replace ARS with $
            proof_number_pattern = r'\b\d{4}\b'
            payer_name_pattern = "None"
            cuit_pattern = "None"  # MARTIN SAID Cuit IS FILLED MANUALLY

            # Extract information using the regEx patterns
            details_regEx_patterns()

            # bank = "HSBC"
            dated = dateds_found[0] if dateds_found else None
            amount = amounts_found[0] if amounts_found else None
            amount = amount.replace('ARS', '$ ') #replaces '$' with '$ '
            amount = amount.replace('.', '')
            amount = amount.split(',')[0] #cuts off upto ,
            payer = payer_name_found[0] if payer_name_found else None
            cuit = cuit_found[0] if cuit_found else None
            proof_number = proof_number_found[1] if proof_number_found else None

    elif bank_name == "Uala":

        # for  uala.png and uala1.png formats
        if lines[0] == "afr":
            # global line_8, line_10
            #payer name
            line_8 = lines[9]
            line_8 = line_8.replace("Nombre remitente ", "")
            # print(line_8)
            #for proof number
            line_10 = lines[11]
            line_10 = line_10.replace("Id Op. ", "")
            # print(line_10)

            #bank_pattern = 'Uala'
            dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
            amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
            proof_number_pattern = line_10
            payer_name_pattern = line_8
            cuit_pattern = "None"  # MARTIN SAID Cuit IS FILLED MANUALLY

            # Extract information using the regEx patterns
            details_regEx_patterns()

            #bank = bank_found[0] if bank_found else None
            dated = dateds_found[0] if dateds_found else None
            amount = amounts_found[0] if amounts_found else None
            amount = amount.replace('$', '$ ') #replaces '$' with '$ '
            amount = amount.replace('.', '')
            payer = payer_name_found[0] if payer_name_found else None
            cuit = cuit_found[0] if cuit_found else None
            proof_number = proof_number_found[0] if proof_number_found else None

        # if lines[0] == "VAD Comprobante de transferencia":
        else:
            #payer name
            line_8 = lines[8]
            line_8 = line_8.replace("Nombre remitente ", "")
            # print(line_8)
            #for proof number
            line_10 = lines[10]
            line_10 = line_10.replace("Id Op. ", "")

            #bank_pattern = 'Uala'
            dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
            amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
            proof_number_pattern = line_10
            payer_name_pattern = line_8
            cuit_pattern = "None"  # MARTIN SAID Cuit IS FILLED MANUALLY

            # Extract information using the regEx patterns
            details_regEx_patterns()

            #bank = bank_found[0] if bank_found else None
            dated = dateds_found[0] if dateds_found else None
            amount = amounts_found[0] if amounts_found else None
            amount = amount.replace('$', '$ ') #replaces '$' with '$ '
            amount = amount.replace('.', '')
            payer = payer_name_found[0] if payer_name_found else None
            cuit = cuit_found[0] if cuit_found else None
            proof_number = proof_number_found[0] if proof_number_found else None

        #one error, language not supported 
    elif bank_name == "Santander":
        #bank_pattern = 'Santander'
        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = 'None' #MARTIN SAID NAME IS FILLED MANUALLY 
        cuit_pattern = 'None' #MARTIN SAID Cuit IS FILLED MANUALLY 

        # Extract information using the regEx patterns
        details_regEx_patterns()

            #bank = bank_found[0] if bank_found else None
        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "Cuenta DNI":

        #bank_pattern = 'Cuenta DNI'
        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        cuit_pattern = r'\b\d{11}\b'

        # if lines[0] == "fo al":
        # if lines[1] == "v4 DNI":
            #payer name
        lines_10 = lines[10]
            
        proof_number_pattern = r'\b\d{6,8}\b' #6,8
        payer_name_pattern = lines_10  

        # Extract information using the regEx patterns
        details_regEx_patterns()

        # #bank = bank_found[0] if bank_found else None
        #bank = #bank_pattern
        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[0] if amounts_found else None
        amount = amount.replace('.', '')
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    elif bank_name == "Macro":

        #for payer name
        line_14 = lines[14]
        
        #bank_pattern = 'Macro' #'usec' #HSBC is detected as that
        dated_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern =  r'\$\s*\d+\,?\d*' 
        #r'\$\s*\d[\d,]*' #r'\$\s*\d+\.?\d*' #r'\$\s*\d[\d,\.]*'
        proof_number_pattern = r'\b\d{9}\b'
        payer_name_pattern = line_14
        cuit_pattern = "None" #because it's not detecting correctly #r"\b\d{11}\b"

        # Extract information using the regEx patterns
        details_regEx_patterns()
        
        #bank = "Macro"
        dated = dateds_found[0] if dateds_found else None
        amount = amounts_found[1] if amounts_found else None
        amount = re.sub(',', '', amount) #amount.replace(',', '')
        amount = amount.split('.')[0]
        payer = payer_name_found[0] if payer_name_found else None
        cuit = cuit_found[0] if cuit_found else None
        proof_number = proof_number_found[0] if proof_number_found else None

    else:
        move_to_undetected(image_path, folder_path)
    # Return the extracted data
    return {
        'BANCO': bank_name,
        'FECHA': dated,
        'IMPORTE': amount,
        'NRO COMPROBANTE': proof_number,
        'TITULAR': payer,
        'CUIT': cuit
    }

# Function to check if a cell is empty
def is_empty(cell_value):
    return cell_value is None or cell_value == ''

# Function to check images in the folder and add data to Excel sheet
def check_image_and_padding(folder_path):
    global index, last_amt
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for filename in os.listdir(folder_path):
        if filename.lower().endswith(('.jpg', '.jpeg', '.png')) and filename not in processed_files:
            processed_files.add(filename)
            # Extract data from the image
            image_path = os.path.join(folder_path, filename)
            extracted_data = extract_data_from_image(image_path)
            if extracted_data:
                # Add data to Excel sheet
                row_data = [index]
                row_data.extend([extracted_data.get(header, '') for header in headers[1:]])
                ws.append(row_data)
                index += 1
                            # Updated total sum formula dynamically
                if extracted_data['IMPORTE']:
                    amt = re.sub(r'[^\$\s0-9.]', '', extracted_data['IMPORTE'])
                    amt = re.sub(r'[^\d.,]', '', amt) #r'[^\d.]'
                    num_amt = Decimal(amt)
                    # total_sum_formula += num_amt

                    last_amt.append(num_amt)
                    print(f"Current total sum: {sum(last_amt)}")

                # Apply yellow fill to empty cells
                if any(is_empty(cell.value) for cell in ws[ws.max_row]):
                    for cell in ws[ws.max_row]:
                        if is_empty(cell.value):
                            cell.fill = yellow_fill
    # Adjust column widths
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

# Process images in the folder and save Excel file
pdf_to_image(folder_path)

print(' ')
for char in ['P','D','F','s',' ','c','o','n','v','e','r','t','e','d',' ','t','o',' ','i','ma','g','e','s',' ','s','u','c','c','e','s','s','f','u','l','l','y','.','.','.','.']:
# for char in b:
    print(char, end='')
    time.sleep(0.1)
print('')

print('')

for char in ['v','e','r','i','f','y','i','n','g',' ','t','h','e',' ','i','m','a','g','e','s','.','.','.']:
    print(char, end="")
    time.sleep(0.1)
print()

chk_anim = ['checking','checking.','checking..','checking...','checking....']
for i in range(3):
    for char in chk_anim:
        print(char, end="\r")
        time.sleep(1)
    print(' ' * len(chk_anim[-1]), end="\r")  # Erase the line
    time.sleep(0.5)

print()
for char in ['c','h','e','c','k','i','n','g',' ','c','o','m','p','l','e','t','e','d','.','.','.',' ','m','o','v','i','n','g',' ','o','n',' ','t','o',' ','e','x','t','r','a','c','t','i','o','n']:    
    print(char, end='')
    time.sleep(0.05)
print("\n")
print(' ')

check_image_and_padding(folder_path)


# sum and adding and saving and opening the file

total_sum_row = [None, None, None, None, None, None, None]
ws.append(total_sum_row)

total_sum = sum(last_amt)
# Append the total sum row to the Excel sheet
total_sum_row = ['SUMA TOTAL', None, f'${total_sum}', None, None, None, None]
ws.append(total_sum_row)

from datetime import date

today = date.today()
formatted_dated = today.strftime("%d_%m_%Y")

# Specify the path to the extracted file
extracted_file_path = os.path.join(folder_path, 'detected')
os.makedirs(extracted_file_path, exist_ok=True)
extracted_file_path = os.path.join(extracted_file_path, f'{formatted_dated}_extracted_info.xlsx')


# Save the Excel file
wb.save(extracted_file_path)


# subprocess.run(['libreoffice', extracted_file_path])
# Open the file with Microsoft Excel in windows
os.system(f'start "{extracted_file_path}"')
