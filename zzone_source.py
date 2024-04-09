import cv2
import pytesseract
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from decimal import Decimal
import re


# Global variables
index = 1
last_amt = []
wb = Workbook()
ws = wb.active
ws.title = "Extracted Data"
headers = ['Serie', 'FECHA', 'IMPORTE', 'NRO COMPROBANTE', 'TITULAR', 'CUIT', 'BANCO']
ws.append(headers)
bold_font = Font(bold=True)
for cell in ws[1]:
    cell.font = bold_font
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def pdf_to_image(pdf_folder_path):
    pdf_files = [file for file in os.listdir(pdf_folder_path) if file.endswith('.pdf')]
    for pdf_file in pdf_files:
        pdf_file_path = os.path.join(pdf_folder_path, pdf_file)
        image_file_path = os.path.join(pdf_folder_path, os.path.splitext(pdf_file)[0] + '.png')
        images = convert_from_path(pdf_file_path)
        for i, image in enumerate(images):
            image.save(image_file_path, 'PNG')
    print('PDFs converted to images successfully.')


def extract_image_to_text(image_path):
    global extracted_text
    image = cv2.imread(image_path)
    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    extracted_text = pytesseract.image_to_string(gray_image)


def details_regEx_patterns():
    global bank_found, dates_found, amounts_found, payer_name_found, cuit_found, proof_number_found
    bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
    dates_found = re.findall(date_pattern, extracted_text)
    amounts_found = re.findall(amount_pattern, extracted_text)
    payer_name_found = re.findall(payer_name_pattern, extracted_text)
    cuit_found = re.findall(cuit_pattern, extracted_text)
    proof_number_found = re.findall(proof_number_pattern, extracted_text)
    return bank_found, dates_found, amounts_found, payer_name_found, cuit_found, proof_number_found


def extract_details():
    global bank, date, amount, payer, cuit, proof_number
    bank = bank_found[0] if bank_found else None
    date = dates_found[0] if dates_found else None
    amount = amounts_found[0] if amounts_found else None
    proof_number = proof_number_found[0] if proof_number_found else None
    payer = payer_name_found[0] if payer_name_found else None
    cuit = cuit_found[0] if cuit_found else None


def get_extracted_details(bank, date, amount, payer, cuit, proof_number):
    return {
        'BANCO': bank,
        'FECHA': date,
        'IMPORTE': amount,
        'TITULAR': payer,
        'CUIT': cuit,
        'NRO COMPROBANTE': proof_number
    }


def check_image_and_padding(folder_path):
    global index, last_amt, ws
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for filename in os.listdir(folder_path):
        if filename.endswith(".jpg") or filename.endswith(".png") or filename.endswith(".jpeg"):
            image_path = os.path.join(folder_path, filename)
            extracted_data = extract_data_from_image(image_path)
            row_data = [index]
            row_data.extend([extracted_data.get(header) for header in headers[1:]])
            ws.append(row_data)
            index += 1
            if extracted_data['IMPORTE']:
                amt = re.sub(r'[^\$\s0-9.]', '', extracted_data['IMPORTE'])
                num_amt = Decimal(re.sub(r'[^\d.]', '', amt))
                last_amt.append(num_amt)
                print(f"Current total sum: {sum(last_amt)}")
            if any(is_empty(cell.value) for cell in ws[ws.max_row]):
                for cell in ws[ws.max_row]:
                    if is_empty(cell.value):
                        cell.fill = yellow_fill
    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width


def extract_data_from_image(image_path):
    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    extract_image_to_text(image_path)

    lines = extracted_text.strip().split('\n')

    if any("BANCOPATAGONIA" in line.lower() for line in lines):
        # global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern
        lines_11 = lines[11]
        bank_pattern = 'BANCOPATAGONIA'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*'
        proof_number_pattern = r'\b\d{10}\b'
        payer_name_pattern = lines_11
        cuit_pattern = 'None'

    if any("galicia" in line.lower() for line in lines):
        # global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern
        line_8 = lines[8]
        bank_pattern = 'galicia'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*'
        proof_number_pattern = r'\b\d{9,11}\b'
        payer_name_pattern = line_8
        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

    if any("mercado pago" in line.lower() for line in lines):
        # global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern
        bank_pattern = 'mercado pago'
        date_pattern = r'\b\d{1,2} de [a-z]+ \d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*'
        payer_name_pattern = r'(?:de )?([A-Z][a-z]+(?: [A-Z][a-z]+)*)'
        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'
        proof_number_pattern = r'\b\d{11}\b'

    if any("santander" in line.lower() for line in lines):
        # global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern
        bank_pattern = 'santander'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = 'None'
        cuit_pattern = 'None'

    if any("supervielle" in line.lower() for line in lines):
        # global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern
        bank_pattern = 'supervielle'
        date_pattern = r'\b\d{1,2}\s+[A-Za-z]+\s+\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*'
        proof_number_pattern = r'\b\d{4}\b'
        payer_name_pattern = 'None' #MARTIN said NAME IS FILLED MANUALLY 
        cuit_pattern =  'None' #MARTIN said Cuit IS FILLED MANUALLY 

    if any("bna" in line.lower() for line in lines):
        # global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern
        bank_pattern = 'bna'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*'
        proof_number_pattern = r'\b\d{8}\b' #have to ask martin about this coz there's none
        payer_name_pattern = 'None' #MARTIN said NAME IS FILLED MANUALLY 
        cuit_pattern =  r'\b\d{11}\b'  #MARTIN said Cuit IS FILLED MANUALLY 
    
    """ currently bank name is not dectecting
    if any("cuenta dni" in line.lower() for line in lines):
        global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern
        lines_7 = lines[7]
        bank_pattern = 'Cuenta DNI'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*'
        proof_number_pattern = r'\b\d{6}\b' #have to ask martin about this coz there's none

        payer_name_pattern = lines_7

        cuit_pattern =  r'\b\d{11}\b' """
    
    if any("BancoCiudad" in line.lower() for line in lines):
        # global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern
        line_33 = lines[33] 
        line_34 = line_33 + lines[34]

        bank_pattern = 'BancoCiudad'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = line_34 
        cuit_pattern =  r'\b\d{2}-\d{8}-\d{1}\b'

    if any("Banco Santa Fe" in line.lower() for line in lines):
        # global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern
        line_24 = lines[24] 
        bank_pattern = 'Banco Santa Fe'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = line_24 
        cuit_pattern =  r'\b\d{11}\b'

    details_regEx_patterns()

    extract_details()
    
    if any("bancopatagonia" in line.lower() for line in lines):
        global amount
        amount = amounts_found[1] if amounts_found else None
    
    if any("galicia" in line.lower() for line in lines):
        global proof_number
        if any(len(p) == 9 for p in proof_number_found) and any(len(p) == 11 for p in proof_number_found):
            # Extract the 9-digit proof number if both are present
            proof_number = next((p for p in proof_number_found if len(p) == 9), None)
        else:
            # Extract the first found proof number (either 9-digit or 11-digit)
            proof_number = proof_number_found[0] if proof_number_found else None

    if any("mercado pago" in line.lower() for line in lines):
        global payer
        #this is important to extract the [1] amount coz the fun has [0] default
        payer = payer_name_found[1] if payer_name_found else None

#    if any("cuenta dni" in line.lower() for line in lines):
#        global bank
#        bank = 'Cuenta DNI'

    if any("BancoCiudad" in line.lower() for line in lines):
        global cuit
        #this line is important to extract the proof number
        proof_number = proof_number_found[1] if proof_number_found else None
        #this line is important to extract the payer name
        payer = re.sub(r'[0-9,-]+', '', payer_name_pattern)
        #this line is important to extract the cuit number
        cuit = cuit_found[1] if cuit_found else None

    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)


def is_empty(value):
    return value is None or value == ''


# Example usage
folder_path = 'source'
check_image_and_padding(folder_path)
wb.save('extracted_data.xlsx')



'''import cv2
import pytesseract
import os
from pdf2image import convert_from_path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from decimal import Decimal
import re

# Initialize global variables
index = 1
last_amt = []
wb = Workbook()
ws = wb.active
ws.title = "Extracted Data"
headers = ['Serie', 'FECHA', 'IMPORTE', 'NRO COMPROBANTE', 'TITULAR', 'CUIT', 'BANCO']
ws.append(headers)
bold_font = Font(bold=True)
for cell in ws[1]:
    cell.font = bold_font
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def pdf_to_image(pdf_folder_path):
    pdf_files = [file for file in os.listdir(pdf_folder_path) if file.endswith('.pdf')]
    for pdf_file in pdf_files:
        pdf_file_path = os.path.join(pdf_folder_path, pdf_file)
        image_file_path = os.path.join(pdf_folder_path, os.path.splitext(pdf_file)[0] + '.png')
        images = convert_from_path(pdf_file_path)
        for i, image in enumerate(images):
            image.save(image_file_path, 'PNG')
    print('PDFs converted to images successfully.')

def extract_image_to_text(image_path):
    global extracted_text
    image = cv2.imread(image_path)
    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    extracted_text = pytesseract.image_to_string(gray_image)

def pattern_matcher(bank_name):
    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern, extracted_text

    lines = extracted_text.strip().split('\n')
    
    # if bank_name == 'bancopatagonia':
    #     bank_pattern = 'bancopatagonia'
    #     date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    #     amount_pattern = r'\$\s*\d+\.?\d*'
    #     proof_number_pattern = r'\b\d{10}\b'
    #     payer_name_pattern = extracted_text.split('\n')[11]
    #     cuit_pattern = 'None'
    
    # if bank_name == 'galicia':
    #     bank_pattern = 'galicia'
    #     date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    #     amount_pattern = r'\$\s*\d+\.?\d*'
    #     proof_number_pattern = r'\b\d{9,11}\b'
    #     payer_name_pattern = extracted_text.split('\n')[8]
    #     cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'
    
    # if any("mercado pago" in line.lower() for line in lines):
    #     bank_pattern = 'mercado pago'
    #     date_pattern = r'\b\d{1,2} de [a-z]+ \d{4}\b'
    #     amount_pattern = r'\$\s*\d+\.?\d*'
    #     payer_name_pattern = r'(?:de )?([A-Z][a-z]+(?: [A-Z][a-z]+)*)'
    #     cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'
    #     proof_number_pattern = r'\b\d{11}\b'

    # if any("santander" in line.lower() for line in lines):
    #     bank_pattern = 'santander'
    #     date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    #     amount_pattern = r'\$\s*\d+\.?\d*'
    #     proof_number_pattern = r'\b\d{8}\b'
    #     payer_name_pattern = 'None'
    #     cuit_pattern = 'None'

    # if any("supervielle" in line.lower() for line in lines):
    #     bank_pattern = 'supervielle'
    #     date_pattern = r'\b\d{1,2}\s+[A-Za-z]+\s+\d{4}\b'
    #     amount_pattern = r'\$\s*\d+\.?\d*'
    #     proof_number_pattern = r'\b\d{4}\b'
    #     payer_name_pattern = 'None'
    #     cuit_pattern = 'None'

    # if any("bna" in line.lower() for line in lines):
    #     bank_pattern = 'bna'
    #     date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    #     amount_pattern = r'\$\s*\d+\.?\d*'
    #     proof_number_pattern = r'\b\d{8}\b'
    #     payer_name_pattern = 'None'
    #     cuit_pattern = r'\b\d{11}\b'

    # if any("BancoCiudad" in line.lower() for line in lines):
    #     bank_pattern = 'BancoCiudad'
    #     date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    #     amount_pattern = r'\$\s*\d+\.?\d*'
    #     proof_number_pattern = r'\b\d{8}\b'
    #     payer_name_pattern = lines[34]
    #     cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

    # if any("Banco Santa Fe" in line.lower() for line in lines):
    #     bank_pattern = 'Banco Santa Fe'
    #     date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
    #     amount_pattern = r'\$\s*\d+\.?\d*'
    #     proof_number_pattern = r'\b\d{8}\b'
    #     payer_name_pattern = lines[24]
    #     cuit_pattern = r'\b\d{11}\b'

    # details_regEx_patterns()

    #return bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

def details_regEx_patterns():
    global bank_found, dates_found, amounts_found, payer_name_found, cuit_found, proof_number_found, extracted_text
    
    lines = extracted_text.strip().split('\n')
    bank_found = re.findall(bank_pattern, extracted_text, re.IGNORECASE)
    dates_found = re.findall(date_pattern, extracted_text)
    amounts_found = re.findall(amount_pattern, extracted_text)
    payer_name_found = re.findall(payer_name_pattern, extracted_text)
    cuit_found = re.findall(cuit_pattern, extracted_text)
    proof_number_found = re.findall(proof_number_pattern, extracted_text)
    
    extract_details()

    # if any("bancopatagonia" in line.lower() for line in lines):
    #     global amount
    #     amount = amounts_found[1] if amounts_found else None

    # if any("galicia" in line.lower() for line in lines):
    #     global proof_number
    #     if any(len(p) == 9 for p in proof_number_found) and any(len(p) == 11 for p in proof_number_found):
    #         proof_number = next((p for p in proof_number_found if len(p) == 9), None)
    #     else:
    #         proof_number = proof_number_found[0] if proof_number_found else None
    
    # if any("mercado pago" in line.lower() for line in lines):
    #     global payer
    #     payer = payer_name_found[1] if payer_name_found else None

    # if any("BancoCiudad" in line.lower() for line in lines):
    #     global cuit
    #     proof_number = proof_number_found[1] if proof_number_found else None
    #     payer = re.sub(r'[0-9,-]+', '', payer_name_pattern)
    #     cuit = cuit_found[1] if cuit_found else None
    
    # return get_extracted_details(bank, date, amount, payer, cuit, proof_number)

    #return bank_found, dates_found, amounts_found, payer_name_found, cuit_found, proof_number_found


def extract_details():
    global bank, date, amount, payer, cuit, proof_number, bank_found, dates_found, amounts_found, payer_name_found, cuit_found, proof_number_found

    bank = bank_found[0] if bank_found else None
    date = dates_found[0] if dates_found else None
    amount = amounts_found[0] if amounts_found else None
    proof_number = proof_number_found[0] if proof_number_found else None
    payer = payer_name_found[0] if payer_name_found else None
    cuit = cuit_found[0] if cuit_found else None


def get_extracted_details(bank, date, amount, payer, cuit, proof_number):
    return {
        'BANCO': bank,
        'FECHA': date,
        'IMPORTE': amount,
        'TITULAR': payer,
        'CUIT': cuit,
        'NRO COMPROBANTE': proof_number
    }


def check_image_and_padding(folder_path):
    global index, last_amt, ws
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for filename in os.listdir(folder_path):
        if filename.endswith(".jpg") or filename.endswith(".png") or filename.endswith(".jpeg"):
            image_path = os.path.join(folder_path, filename)
            extracted_data = extract_data_from_image(image_path)
            if extracted_data is not None:
                row_data = [index]
                row_data.extend([extracted_data.get(header) for header in headers[1:]])
                ws.append(row_data)
                index += 1
                if extracted_data['IMPORTE']:
                    amt = re.sub(r'[^\$\s0-9.]', '', extracted_data['IMPORTE'])
                    num_amt = Decimal(re.sub(r'[^\d.]', '', amt))
                    last_amt.append(num_amt)
                    print(f"Current total sum: {sum(last_amt)}")
                if any(is_empty(cell.value) for cell in ws[ws.max_row]):
                    for cell in ws[ws.max_row]:
                        if is_empty(cell.value):
                            cell.fill = yellow_fill
    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width


def extract_data_from_image(image_path):
    global bank_pattern, date_pattern, amount_pattern, proof_number_pattern, payer_name_pattern, cuit_pattern

    extract_image_to_text(image_path)

    lines = extracted_text.strip().split('\n')

    line = lines[0].lower()
    if line == "bancopatagonia":
        pattern_matcher('bancopatagonia')
    elif line == "galicia":
        pattern_matcher('galicia')
    """
    line = lines[0].lower()

    #if line == "bancopatagonia":
    if any("bancopatagonia" in line.lower() for line in lines):
        bank_pattern = 'bancopatagonia'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*'
        proof_number_pattern = r'\b\d{10}\b'
        payer_name_pattern = lines[11]
        cuit_pattern = 'None'
    
    if any("galicia" in line.lower() for line in lines):
        bank_pattern = 'galicia'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*'
        proof_number_pattern = r'\b\d{9,11}\b'
        payer_name_pattern = lines[8]
        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

    if any("mercado pago" in line.lower() for line in lines):
        bank_pattern = 'mercado pago'
        date_pattern = r'\b\d{1,2} de [a-z]+ \d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*'
        payer_name_pattern = r'(?:de )?([A-Z][a-z]+(?: [A-Z][a-z]+)*)'
        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'
        proof_number_pattern = r'\b\d{11}\b'

    if any("santander" in line.lower() for line in lines):
        bank_pattern = 'santander'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = 'None'
        cuit_pattern = 'None'

    if any("supervielle" in line.lower() for line in lines):
        bank_pattern = 'supervielle'
        date_pattern = r'\b\d{1,2}\s+[A-Za-z]+\s+\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*'
        proof_number_pattern = r'\b\d{4}\b'
        payer_name_pattern = 'None'
        cuit_pattern = 'None'

    if any("bna" in line.lower() for line in lines):
        bank_pattern = 'bna'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = 'None'
        cuit_pattern = r'\b\d{11}\b'

    if any("BancoCiudad" in line.lower() for line in lines):
        bank_pattern = 'BancoCiudad'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = lines[34]
        cuit_pattern = r'\b\d{2}-\d{8}-\d{1}\b'

    if any("Banco Santa Fe" in line.lower() for line in lines):
        bank_pattern = 'Banco Santa Fe'
        date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b'
        amount_pattern = r'\$\s*\d+\.?\d*'
        proof_number_pattern = r'\b\d{8}\b'
        payer_name_pattern = lines[24]
        cuit_pattern = r'\b\d{11}\b'
    """
    """details_regEx_patterns()"""
    """extract_details()"""

    """if any("bancopatagonia" in line.lower() for line in lines):
        global amount
        amount = amounts_found[1] if amounts_found else None

    if any("galicia" in line.lower() for line in lines):
        global proof_number
        if any(len(p) == 9 for p in proof_number_found) and any(len(p) == 11 for p in proof_number_found):
            proof_number = next((p for p in proof_number_found if len(p) == 9), None)
        else:
            proof_number = proof_number_found[0] if proof_number_found else None
    
    if any("mercado pago" in line.lower() for line in lines):
        global payer
        payer = payer_name_found[1] if payer_name_found else None

    if any("BancoCiudad" in line.lower() for line in lines):
        global cuit
        proof_number = proof_number_found[1] if proof_number_found else None
        payer = re.sub(r'[0-9,-]+', '', payer_name_pattern)
        cuit = cuit_found[1] if cuit_found else None
    
    return get_extracted_details(bank, date, amount, payer, cuit, proof_number)"""


def is_empty(value):
    return value is None or value == ''


# Example usage
# folder_path = 'bancopatagonia/bancopatagonia_assets'
# check_image_and_padding(folder_path)
# folder_path = 'galicia/galicia_assets'
folder_path = 'source'
check_image_and_padding(folder_path)
wb.save('extracted_data.xlsx')'''
